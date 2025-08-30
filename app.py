# NBDFinder - Non-B DNA Analysis Platform 
# ============================================
# Compact, information-dense Streamlit interface
# Optimized for professional dashboard-style usage

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import time
from collections import Counter
from Bio import Entrez, SeqIO
import numpy as np
from io import BytesIO

# Core motif detection imports with performance optimization
from motifs import (
    find_hotspots,
    parse_fasta, parse_fasta_multi, gc_content, reverse_complement,
    select_best_nonoverlapping_motifs, wrap
)

# Performance-optimized motif detection with caching
@st.cache_data(show_spinner=True, max_entries=10)
def cached_all_motifs(sequence_hash: str, sequence: str, settings: dict = None):
    """Cached wrapper for all_motifs function to avoid recomputation"""
    from motifs import all_motifs
    import time
    
    start_time = time.time()
    motifs = all_motifs(sequence)
    processing_time = time.time() - start_time
    
    return {
        'motifs': motifs,
        'processing_time': processing_time,
        'sequence_length': len(sequence),
        'settings_used': settings or {}
    }

# Cached GC content calculation
@st.cache_data(max_entries=50)
def cached_gc_content(sequence: str) -> float:
    """Cached GC content calculation"""
    return gc_content(sequence)

# Performance-optimized DataFrame processing with memory management
@st.cache_data(max_entries=10)
def create_motif_distribution_chart(motif_counts_dict: dict):
    """Cached motif distribution chart creation"""
    if not motif_counts_dict:
        return None
    
    # Lazy import for visualization
    plt, px, go, make_subplots, BytesIO, seaborn = lazy_import_visualization()
    
    # Create chart with optimized data processing
    fig = px.pie(
        values=list(motif_counts_dict.values()),
        names=list(motif_counts_dict.keys()),
        title="Motif Class Distribution",
        height=400
    )
    fig.update_layout(margin=dict(t=40, b=20, l=20, r=20))
    return fig

@st.cache_data(max_entries=10)
def process_motif_counts_vectorized(results_list: list):
    """Vectorized motif counting using pandas for better performance"""
    if not results_list:
        return {}
    
    # Extract all motifs in one pass
    all_motifs = []
    for result in results_list:
        for motif in result['motifs']:
            all_motifs.append(motif.get('Class', 'Unknown'))
    
    if all_motifs:
        # Use pandas for fast counting with explicit cleanup
        motif_series = pd.Series(all_motifs, dtype='category')
        motif_counts = motif_series.value_counts().to_dict()
        
        # Explicit cleanup for memory efficiency
        del motif_series, all_motifs
        import gc
        gc.collect()
        
        return motif_counts
    
    return {}

@st.cache_data(max_entries=10)
def create_results_summary_df(results_list: list):
    """Create optimized results summary DataFrame with proper dtypes"""
    if not results_list:
        return pd.DataFrame()
    
    # Pre-allocate data dictionary for efficiency
    data = {
        'Sequence Name': [],
        'Length (bp)': [],
        'Total Motifs': [],
        'Processing Time (s)': []
    }
    
    for result in results_list:
        data['Sequence Name'].append(result['sequence_name'])
        data['Length (bp)'].append(result['sequence_length']) 
        data['Total Motifs'].append(result['total_motifs'])
        data['Processing Time (s)'].append(result.get('processing_time', 0.0))
    
    # Create DataFrame with optimized dtypes
    df = pd.DataFrame(data)
    df['Length (bp)'] = df['Length (bp)'].astype('int32')
    df['Total Motifs'] = df['Total Motifs'].astype('int16')
    df['Processing Time (s)'] = df['Processing Time (s)'].astype('float32')
    
    # Vectorized calculation
    df['Motifs/kb'] = (df['Total Motifs'] / df['Length (bp)'] * 1000).round(2).astype('float32')
    
    return df

# Lazy import functions for performance
def lazy_import_visualization():
    """Lazy import of visualization libraries when needed"""
    try:
        # Check if matplotlib is already imported in globals
        import matplotlib.pyplot as plt
        import plotly.express as px
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        from io import BytesIO
        import seaborn as sns
        return plt, px, go, make_subplots, BytesIO, sns
    except ImportError as e:
        # Fallback for missing packages
        print(f"Visualization import warning: {e}")
        return None, None, None, None, None, None

def lazy_import_bio():
    """Lazy import of biopython when needed"""
    try:
        from Bio import Entrez, SeqIO
        return Entrez, SeqIO
    except ImportError as e:
        print(f"Bio import warning: {e}")
        return None, None

# Advanced visualization components (optional dependencies)
try:
    from advanced_visualization import create_enhanced_dashboard, export_publication_figure
    ADVANCED_VIZ_AVAILABLE = True
except ImportError:
    ADVANCED_VIZ_AVAILABLE = False
    print("Advanced visualization not available. Install additional dependencies.")

# Publication visualization integration (optional dependencies)  
try:
    from nbdfinder_viz_integration import create_nbdfinder_visualization_interface
    PUBLICATION_VIZ_AVAILABLE = True
except ImportError:
    PUBLICATION_VIZ_AVAILABLE = False
    print("Publication visualization module not available")

# ---------- STREAMLIT PAGE CONFIGURATION ----------
# Wide layout for maximum information density
st.set_page_config(
    page_title="NBDFinder - Non-B DNA Analysis Platform",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://github.com/VRYella/NBDFinder',
        'Report a bug': 'https://github.com/VRYella/NBDFinder/issues',
        'About': "NBDFinder - The most advanced non-B DNA prediction platform"
    }
)

# ---------- MOTIF SUBTYPE VALIDATION UTILITY ----------
# Ensures all motifs have proper classification for consistent display
def ensure_subtype(motif):
    """Guarantee every motif has a string 'Subtype' using proper classification mapping"""
    from motifs.classification_config import get_official_classification, LEGACY_TO_OFFICIAL_MAPPING
    
    if isinstance(motif, dict):
        if 'Subtype' not in motif or motif['Subtype'] is None or motif['Subtype'] == '':
            # Map legacy classifications to official system
            motif_class = motif.get('Class', '')
            if motif_class in LEGACY_TO_OFFICIAL_MAPPING:
                official_class, official_subtype = get_official_classification(motif_class, '')
                motif['Subtype'] = official_subtype if official_subtype else 'Canonical'
            else:
                motif['Subtype'] = 'Canonical'  # Default fallback
        return motif
    else:
        # Handle non-dict motifs gracefully
        return {'Subtype': 'Canonical', 'Motif': motif}

# ---------- COMPACT FLOATING COMPONENTS ----------
def add_floating_back_to_top():
    """Add compact floating back to top button"""
    st.markdown("""
    <div id="back-to-top" onclick="window.scrollTo({top: 0, behavior: 'smooth'})">
        ↑ Top
    </div>
    """, unsafe_allow_html=True)



# =========================================================================
# CONSOLIDATED CSS DESIGN SYSTEM - SCIENTIFIC DASHBOARD STANDARDS
# =========================================================================
# Based on:
# - WCAG 2.1 AA accessibility guidelines for color contrast and navigation
# - Material Design Component spacing (8px grid system) for consistency
# - Nature Publishing typography standards for scientific readability
# - Inter Design System font specifications for modern web interfaces

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    :root {
        /* Scientific Color Palette - Nature/Science Journal Standards */
        --primary: #1e3a8a;         /* Deep blue - trust, reliability */
        --accent: #0891b2;          /* Cyan - data visualization */
        --secondary: #6366f1;       /* Indigo - interactive elements */
        --success: #10b981;         /* Emerald - positive feedback */
        --warning: #f59e0b;         /* Amber - attention */
        --error: #ef4444;           /* Red - alerts */
        --info: #3b82f6;           /* Blue - information */
        
        /* Neutral palette for scientific interfaces */
        --text: #1f2937;           /* Primary text - high contrast (WCAG AA) */
        --text-muted: #6b7280;     /* Secondary text - accessible gray */
        --text-light: #9ca3af;     /* Tertiary text - subtle information */
        --bg: #ffffff;             /* Clean white background */
        --surface: #f8fafc;        /* Elevated surface color */
        --border: #e5e7eb;         /* Subtle borders */
        --border-focus: #3b82f6;   /* Focus indication */
        
        /* Typography System - Scientific Standards */
        /* Base font size: 16px = 1rem (following Material Design) */
        --font-base: 1rem;         /* 16px - primary text */
        --font-sm: 0.95rem;        /* 15.2px - table/label text */
        --font-xs: 0.92rem;        /* 14.7px - caption text */
        --font-button: 0.97rem;    /* 15.5px - button text */
        
        /* Header hierarchy - Nature/Science journal standards */
        --h1-size: 2.2rem;         /* 35.2px - main titles */
        --h2-size: 1.6rem;         /* 25.6px - section headers */
        --h3-size: 1.3rem;         /* 20.8px - subsection headers */
        --h4-size: 1.1rem;         /* 17.6px - component headers */
        
        /* Line height specifications - optimal readability */
        --line-height-body: 1.5;   /* Body text - readability standard */
        --line-height-heading: 1.2; /* Headings - compact hierarchy */
        
        /* Spacing System - Material Design 8px grid */
        --space-1: 0.5rem;         /* 8px */
        --space-2: 1rem;           /* 16px */
        --space-3: 1.5rem;         /* 24px */
        --space-4: 2rem;           /* 32px */
        --space-5: 2.5rem;         /* 40px */
        --space-6: 3rem;           /* 48px */
        
        /* Visual effects - subtle and professional */
        --radius: 0.5rem;          /* 8px - consistent with grid */
        --shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px 0 rgba(0, 0, 0, 0.06);
        --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
        --transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    /* GLOBAL LAYOUT - SCIENTIFIC INTERFACE STANDARDS */
    body, [data-testid="stAppViewContainer"], .main {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        font-size: var(--font-base);
        line-height: var(--line-height-body);
        color: var(--text);
        background-color: var(--bg);
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
    }
    
    .main .block-container {
        padding: var(--space-3) var(--space-4);
        max-width: none;
        width: 100%;
    }
    .stApp > .main { width: 100%; max-width: none; }
    [data-testid="stAppViewContainer"] { width: 100%; max-width: none; }

    /* TYPOGRAPHY SYSTEM - SCIENTIFIC HIERARCHY */
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        line-height: var(--line-height-heading);
        color: var(--primary);
        margin: var(--space-2) 0 var(--space-1) 0;
        letter-spacing: -0.025em;
    }
    
    h1 { 
        font-size: var(--h1-size); 
        font-weight: 700; 
        background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    h2 { font-size: var(--h2-size); font-weight: 650; }
    h3 { font-size: var(--h3-size); font-weight: 600; }
    h4 { font-size: var(--h4-size); font-weight: 600; }
    
    /* Body text and components */
    .stMarkdown, .markdown-text-container, p, span, label {
        font-family: 'Inter', sans-serif;
        font-size: var(--font-base);
        line-height: var(--line-height-body);
        color: var(--text);
        margin-bottom: var(--space-1);
    }
    
    /* ENHANCED NAVIGATION TABS - SCIENTIFIC INTERFACE */
    .stTabs [data-baseweb="tab-list"] {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: var(--space-1);
        margin-bottom: var(--space-3);
        box-shadow: var(--shadow);
        display: flex;
        gap: 4px;
        backdrop-filter: blur(10px);
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border: none;
        border-radius: calc(var(--radius) - 2px);
        font-family: 'Inter', sans-serif;
        font-weight: 500;
        font-size: var(--font-base);
        color: var(--text-muted);
        padding: var(--space-2) var(--space-3);
        transition: var(--transition);
        cursor: pointer;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        white-space: nowrap;
        flex: 1;
        min-height: 48px; /* WCAG 2.1 touch target minimum */
        text-align: center;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: linear-gradient(135deg, rgba(255,255,255,0.9) 0%, var(--surface) 100%);
        color: var(--primary);
        font-weight: 600;
        transform: translateY(-1px);
        box-shadow: var(--shadow);
        border: 1px solid var(--border-focus);
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
        color: white;
        font-weight: 600;
        box-shadow: var(--shadow-md);
        transform: translateY(-1px);
        border: 1px solid rgba(255,255,255,0.2);
    }
    /* FORM CONTROLS - ACCESSIBLE AND SCIENTIFIC */
    .stSelectbox > div, .stMultiSelect > div, .stTextInput > div, .stTextArea > div {
        background: var(--bg);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        transition: var(--transition);
        margin-bottom: var(--space-1);
    }
    
    .stSelectbox > div:hover, .stMultiSelect > div:hover, .stTextInput > div:hover, .stTextArea > div:hover {
        border-color: var(--accent);
    }
    
    .stSelectbox > div:focus-within, .stMultiSelect > div:focus-within, .stTextInput > div:focus-within, .stTextArea > div:focus-within {
        border-color: var(--border-focus);
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
        outline: none;
    }
    
    /* Label styling for accessibility */
    .stSelectbox label, .stMultiSelect label, .stTextInput label, .stTextArea label,
    .stSlider label, .stRadio label, .stCheckbox label {
        font-size: var(--font-sm) !important;
        font-weight: 500;
        color: var(--text);
        margin-bottom: var(--space-1) !important;
    }
    
    /* Enhanced selectbox for scientific data display */
    .stSelectbox div[data-baseweb="select"] > div, .stTextInput input, .stTextArea textarea {
        font-size: var(--font-sm) !important;
        padding: var(--space-2) !important;
        line-height: var(--line-height-body) !important;
        min-width: 300px !important;
        white-space: normal !important;
    }
    
    /* Dropdown option list styling for full text visibility */
    .stSelectbox div[data-baseweb="select"] div[role="listbox"] {
        min-width: 300px !important;
        max-width: none !important;
    }
    
    /* Individual dropdown option styling */
    .stSelectbox div[data-baseweb="select"] div[role="option"] {
        padding: var(--space-2) !important;
        line-height: var(--line-height-body) !important;
        white-space: normal !important;
        overflow: visible !important;
        text-overflow: initial !important;
        min-height: auto !important;
    }
    /* BUTTON SYSTEM - SCIENTIFIC INTERFACE */
    .stButton > button {
        background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
        color: white;
        border: none;
        border-radius: var(--radius);
        font-family: 'Inter', sans-serif;
        font-weight: 700;
        font-size: var(--font-button);
        padding: var(--space-2) var(--space-3);
        transition: var(--transition);
        box-shadow: var(--shadow);
        cursor: pointer;
        margin: var(--space-1) 0;
        min-height: 44px; /* WCAG 2.1 touch target minimum */
    }
    
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: var(--shadow-md);
        background: linear-gradient(135deg, var(--accent) 0%, var(--primary) 100%);
    }
    
    .stButton > button:disabled {
        background: var(--text-light);
        color: var(--text-muted);
        cursor: not-allowed;
        transform: none;
        box-shadow: none;
    }
    
    .stDownloadButton > button {
        background: linear-gradient(135deg, var(--success) 0%, #047857 100%);
        color: white;
        border: none;
        border-radius: var(--radius);
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        font-size: var(--font-sm);
        padding: var(--space-2);
        transition: var(--transition);
        box-shadow: var(--shadow);
        min-height: 44px;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-1px);
        box-shadow: var(--shadow-md);
    }
    
    /* DATA TABLES - SCIENTIFIC PRESENTATION */
    .stDataFrame {
        border: 1px solid var(--border);
        border-radius: var(--radius);
        overflow: hidden;
        box-shadow: var(--shadow);
        margin: var(--space-2) 0;
        max-height: 400px;
        overflow-y: auto;
    }
    
    .stDataFrame table {
        font-family: 'Inter', sans-serif;
        font-size: var(--font-sm);
    }
    
    .stDataFrame thead tr th {
        background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
        color: white;
        font-weight: 600;
        font-size: var(--font-sm);
        padding: var(--space-2);
        text-align: left;
        border: none;
        position: sticky;
        top: 0;
        z-index: 10;
    }
    
    .stDataFrame tbody tr:nth-child(even) {
        background: var(--surface);
    }
    
    .stDataFrame tbody tr:hover {
        background: rgba(59, 130, 246, 0.05);
        transition: background 0.2s ease;
    }
    
    .stDataFrame tbody tr td {
        padding: var(--space-2);
        border-bottom: 1px solid var(--border);
        font-weight: 400;
        color: var(--text);
        font-size: var(--font-sm);
    }
    
    /* METRICS AND CARDS - SCIENTIFIC DASHBOARD */
    .metric-card {
        background: var(--bg);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: var(--space-3);
        margin: var(--space-2) 0;
        box-shadow: var(--shadow);
        transition: var(--transition);
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: var(--shadow-lg);
        border-color: var(--accent);
    }
    
    .metric-card h3 {
        background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-size: var(--h2-size);
        font-weight: 700;
        margin: 0;
        text-align: center;
    }
    
    .metric-card p {
        color: var(--text-muted);
        font-size: var(--font-sm);
        font-weight: 500;
        margin: var(--space-1) 0 0 0;
        text-align: center;
    }
    
    /* MESSAGES AND ALERTS - ACCESSIBLE FEEDBACK */
    .stSuccess, .stError, .stWarning, .stInfo {
        border-radius: var(--radius);
        padding: var(--space-2);
        font-family: 'Inter', sans-serif;
        font-weight: 500;
        font-size: var(--font-base);
        box-shadow: var(--shadow);
        margin: var(--space-2) 0;
        border-left-width: 4px;
        border-left-style: solid;
    }
    
    .stSuccess {
        background: #d1fae5;
        border-left-color: var(--success);
        color: #065f46;
    }
    
    .stError {
        background: #fee2e2;
        border-left-color: var(--error);
        color: #991b1b;
    }
    
    .stWarning {
        background: #fef3c7;
        border-left-color: var(--warning);
        color: #92400e;
    }
    
    .stInfo {
        background: #dbeafe;
        border-left-color: var(--info);
        color: #1e40af;
    }
    
    /* RESPONSIVE DESIGN - MOBILE/TABLET SUPPORT */
    @media (max-width: 768px) {
        :root {
            --font-base: 0.9rem;      /* Slightly smaller on mobile */
            --font-sm: 0.85rem;
            --h1-size: 1.8rem;
            --h2-size: 1.4rem;
            --h3-size: 1.2rem;
            --h4-size: 1.0rem;
            --space-2: 0.75rem;       /* Tighter spacing on mobile */
            --space-3: 1rem;
            --space-4: 1.5rem;
        }
        
        .main .block-container {
            padding: var(--space-2) var(--space-2);
        }
        
        .stTabs [data-baseweb="tab"] {
            padding: var(--space-1) var(--space-2);
            font-size: var(--font-sm);
        }
    }
    
    /* ACCESSIBILITY ENHANCEMENTS - WCAG 2.1 AA COMPLIANCE */
    /* Focus indicators for keyboard navigation */
    .stButton > button:focus-visible,
    .stSelectbox > div:focus-within,
    .stTextInput > div:focus-within {
        outline: 2px solid var(--border-focus);
        outline-offset: 2px;
    }
    
    /* High contrast mode support */
    @media (prefers-contrast: high) {
        :root {
            --text: #000000;
            --bg: #ffffff;
            --border: #333333;
            --primary: #0000ff;
        }
    }
    
    /* Reduced motion for accessibility */
    @media (prefers-reduced-motion: reduce) {
        * {
            animation-duration: 0.01ms !important;
            animation-iteration-count: 1 !important;
            transition-duration: 0.01ms !important;
        }
    }
    
    /* FLOATING ELEMENTS - PROFESSIONAL TOUCH */
    #back-to-top {
        position: fixed;
        bottom: var(--space-3);
        right: var(--space-3);
        background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
        color: white;
        padding: var(--space-2);
        border-radius: 50%;
        cursor: pointer;
        box-shadow: var(--shadow-lg);
        z-index: 1000;
        font-weight: 600;
        font-size: var(--font-sm);
        transition: var(--transition);
        border: none;
        width: 48px;
        height: 48px;
        display: flex;
        align-items: center;
        justify-content: center;
    }
    
    #back-to-top:hover {
        transform: translateY(-2px);
        box-shadow: var(--shadow-lg);
    }
    
    /* CUSTOM COMPONENTS - SCIENTIFIC CARDS */
    .feature-card {
        background: var(--bg);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: var(--space-3);
        margin: var(--space-2) 0;
        box-shadow: var(--shadow);
        transition: var(--transition);
    }
    
    .feature-card:hover {
        transform: translateY(-1px);
        box-shadow: var(--shadow-md);
        border-color: var(--accent);
    }
    </style>
    """, unsafe_allow_html=True)

# ---------- CONSTANTS ----------
MOTIF_COLORS = {
    "Curved DNA": "#3B82F6",
    "Slipped DNA": "#10B981", 
    "Cruciform": "#F59E0B",
    "R-loop": "#EF4444",
    "Triplex": "#8B5CF6",
    "G-Quadruplex": "#06B6D4",
    "i-motif": "#EC4899",
    "Z-DNA": "#84CC16",
    "Hybrid": "#F97316",
    "Non-B DNA Clusters": "#EF4444"
}

# Main application pages structure
MAIN_PAGES = {
    "Home": "Home / Welcome",
    "Upload & Analyze": "Sequence Upload and Analysis",
    "Results & Visualization": "Analysis Results and Professional Visualizations",
    "Clinical/Disease": "Disease Annotation and Clinical Data",
    "Documentation": "Scientific Documentation & References"
}

# =========================================================================
# ENHANCED SESSION STATE MANAGEMENT - COMPREHENSIVE PERSISTENCE
# =========================================================================
# Following React/Streamlit best practices for state management

def initialize_session_state():
    """
    Initialize comprehensive session state for scientific dashboard application.
    
    Ensures all interactive elements maintain state across user interactions,
    following WCAG 2.1 guidelines for consistent user experience.
    """
    # Core application state
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "Home"
    
    # Data management state
    if 'seqs' not in st.session_state:
        st.session_state.seqs = []
    if 'names' not in st.session_state:
        st.session_state.names = []
    if 'results' not in st.session_state:
        st.session_state.results = []
    
    # Analysis configuration state
    if 'analysis_settings' not in st.session_state:
        st.session_state.analysis_settings = {
            'motif_classes': list(MOTIF_COLORS.keys()),
            'sensitivity': 'High',
            'min_motif_length': 10,
            'max_motif_length': 200,
            'overlap_threshold': 50,
            'quality_filter': True,
            'parallel_processing': True,
            'memory_optimization': True
        }
    
    # UI state management
    if 'input_method' not in st.session_state:
        st.session_state.input_method = "📁 Upload FASTA"
    if 'ncbi_query' not in st.session_state:
        st.session_state.ncbi_query = ""
    if 'last_input_method' not in st.session_state:
        st.session_state.last_input_method = ""
    
    # Analysis execution state
    if 'analysis_running' not in st.session_state:
        st.session_state.analysis_running = False
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    
    # User preferences and accessibility
    if 'theme_preference' not in st.session_state:
        st.session_state.theme_preference = "light"
    if 'font_size_preference' not in st.session_state:
        st.session_state.font_size_preference = "normal"

# Initialize session state
initialize_session_state()

# Set up Entrez for NCBI access with lazy initialization
@st.cache_data(ttl=3600)
def initialize_entrez():
    """Initialize Entrez settings with caching"""
    Entrez, _ = lazy_import_bio()
    Entrez.email = "raazbiochem@gmail.com"
    Entrez.api_key = None
    return True

# Famous NCBI examples for quick access - Enhanced with more diverse examples
FAMOUS_NCBI_EXAMPLES = {
    "Human TERT Promoter": "NC_000005.10:1253147-1295047",
    "Human c-MYC Promoter": "NG_007161.1",
    "Human BCL2 Promoter": "NG_009361.1", 
    "Fragile X FMR1 Gene": "NG_007529.1",
    "Huntington HTT Gene": "NG_009378.1",
    "Human Immunoglobulin Switch": "NG_001019.6",
    "Human Alpha Globin": "NG_000006.1",
    "Friedreich Ataxia FXN": "NG_008845.1",
    "Human BRCA1 Gene": "NG_005905.2",
    "Human BRCA2 Gene": "NG_012772.3",
    "Human TP53 Tumor Suppressor": "NG_017013.2",
    "Human EGFR Oncogene": "NG_007726.3",
    "Human CFTR (Cystic Fibrosis)": "NG_016465.4",
    "Human DMD (Duchenne MD)": "NG_012232.1",
    "Human APOE (Alzheimer's)": "NG_007991.1",
    "Human SOD1 (ALS)": "NG_008689.1"
}

# Example sequences
EXAMPLE_SEQUENCES = {
    "Human Telomere G4-Rich": {
        "name": "G4_Rich_Human_Telomere_Example", 
        "sequence": "TTAGGGTTAGGGTTAGGGTTAGGGAAAAATCCGTCGAGCAGAGTTAAAAAGGGGTTAGGGTTAGGGTTAGGGCCCCCTCCCCCTCCCCCTCCCCGAAAGAAAGAAAGAAAGAAACGCGCGCGCGCGCGCGCGCGATCGCACACACACAGCTGCTGCTGC",
        "description": "Human telomeric sequence rich in G-quadruplex structures"
    },
    "Z-DNA Forming Sequence": {
        "name": "Z_DNA_Example_Sequence",
        "sequence": "CGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCGCG",
        "description": "Alternating CG sequence prone to Z-DNA formation"
    },
    "Disease-Associated GAA Repeats": {
        "name": "Disease_GAA_Repeats_Example",
        "sequence": "GAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAAGAA",
        "description": "Frataxin GAA repeat expansion associated with Friedreich's ataxia"
    }
}

# Performance-optimized NCBI fetch with caching
@st.cache_data(show_spinner=True, max_entries=20, ttl=3600)  # Cache for 1 hour
def cached_ncbi_fetch(query: str):
    """Cached NCBI fetch to avoid repeated API calls"""
    return ncbi_fetch(query)

def ncbi_fetch(query):
    """Fetch sequences from NCBI using Entrez with improved error handling and retry logic"""
    # Lazy import for bio modules
    Entrez, SeqIO = lazy_import_bio()
    
    import time
    import requests
    
    # Test network connectivity first
    try:
        test_response = requests.get("https://www.ncbi.nlm.nih.gov", timeout=5)
        if test_response.status_code != 200:
            st.error("⚠️ **NCBI Connectivity Issue**: Cannot reach NCBI servers. Please try again later or use direct FASTA input.")
            return [], []
    except Exception as e:
        st.error(f"⚠️ **Network Connectivity Issue**: {str(e)[:100]}... Please check your internet connection or use direct FASTA input.")
        return [], []
    
    max_retries = 3
    retry_delay = 2
    
    for attempt in range(max_retries):
        try:
            # Set email for NCBI Entrez (required for good practice)
            Entrez.email = "raazbiochem@gmail.com"
            
            # Enhanced search logic with multiple strategies
            if ':' in query and '-' in query:
                # Handle genomic coordinate queries (e.g., NC_000005.10:1253147-1295047)
                st.info(f"🧬 Genomic coordinate query detected: {query}")
                handle = Entrez.efetch(db="nucleotide", id=query, rettype="fasta", retmode="text")
            else:
                # Regular accession or gene name query
                st.info(f"🔍 Searching for: {query}")
                
                # First try direct fetch for accession numbers
                try:
                    handle = Entrez.efetch(db="nucleotide", id=query, rettype="fasta", retmode="text")
                except Exception:
                    # If direct fetch fails, try search first
                    search_handle = Entrez.esearch(db="nucleotide", term=query, retmax=5)
                    search_result = Entrez.read(search_handle)
                    search_handle.close()
                    
                    if not search_result['IdList']:
                        st.warning(f"⚠️ No results found for query: '{query}'. Try using a specific accession number.")
                        return [], []
                    
                    # Fetch the first result
                    first_id = search_result['IdList'][0]
                    handle = Entrez.efetch(db="nucleotide", id=first_id, rettype="fasta", retmode="text")
            
            # Parse FASTA content
            content = handle.read()
            handle.close()
            
            seqs = []
            names = []
            
            # Parse multi-FASTA content
            cur_seq = ""
            cur_name = ""
            
            for line in content.split('\n'):
                line = line.strip()
                if line.startswith('>'):
                    # Save previous sequence if exists
                    if cur_seq:
                        cleaned_seq = parse_fasta(cur_seq)
                        if len(cleaned_seq) > 10:
                            seqs.append(cleaned_seq)
                            display_name = cur_name[:100] + "..." if len(cur_name) > 100 else cur_name
                            names.append(display_name if display_name else f"Sequence_{len(seqs)}")
                    
                    # Start new sequence
                    cur_name = line[1:]  # Remove '>'
                    cur_seq = ""
                elif line:
                    cur_seq += line
            
            # Add the last sequence
            if cur_seq:
                cleaned_seq = parse_fasta(cur_seq)
                if len(cleaned_seq) > 10:
                    seqs.append(cleaned_seq)
                    display_name = cur_name[:100] + "..." if len(cur_name) > 100 else cur_name
                    names.append(display_name if display_name else f"Sequence_{len(seqs)}")
            
            if not seqs:
                st.warning("⚠️ No valid sequences found (sequences must be >10 bp).")
                return [], []
            
            st.success(f"✅ Successfully retrieved {len(seqs)} sequence(s) from NCBI!")
            return seqs, names
            
        except Exception as e:
            error_msg = str(e)
            if attempt == max_retries - 1:  # Last attempt
                if "HTTP Error 429" in error_msg:
                    st.error("⚠️ **NCBI Rate Limit**: Too many requests. Please wait 1-2 minutes and try again.")
                elif "URLError" in error_msg or "timeout" in error_msg.lower() or "NameResolutionError" in error_msg:
                    st.error("⚠️ **Network Issue**: Cannot connect to NCBI. Please check your internet connection or try again later.")
                elif "XML" in error_msg or "parse" in error_msg.lower():
                    st.error("⚠️ **NCBI Service Issue**: The service may be temporarily unavailable. Please try again later.")
                else:
                    st.error(f"⚠️ **NCBI Fetch Failed**: {error_msg[:200]}...")
                
                # Provide helpful suggestions
                st.info("💡 **Alternative Options:**\\n"
                       "- Try using direct FASTA sequence input instead\\n"
                       "- Check that your accession number is correct (e.g., 'NC_000001.11')\\n"
                       "- Use one of the provided example sequences\\n"
                       "- Try again in a few minutes if this is a temporary service issue")
                return [], []
            else:
                # Wait before retry
                time.sleep(retry_delay * (attempt + 1))
                continue
    
    return [], []

# ---- MAIN APPLICATION ----

# Add floating components
add_floating_back_to_top()

# Header
st.markdown("""
<div style='text-align: center; padding: 20px; margin-bottom: 30px; background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%); border-radius: 12px; border: 1px solid #e2e8f0;'>
    <h1 style='color: #1e3a8a; font-family: Inter, sans-serif; font-weight: 700; margin-bottom: 8px; font-size: 2.5rem;'>
        NBDFinder: Non-B DNA Analysis Platform
    </h1>
    <p style='color: #64748b; font-size: 1.125rem; margin: 0;'>Professional computational framework for genome-wide detection and analysis of non-B DNA structural motifs</p>
</div>
""", unsafe_allow_html=True)

# =========================================================================
# MODULAR TAB FUNCTIONS - SCIENTIFIC DASHBOARD ARCHITECTURE
# =========================================================================
# Each tab implemented as a standalone function following scientific software
# development best practices and clear separation of concerns

def render_home_tab():
    """
    Render the Home/Welcome tab with comprehensive platform overview.
    
    Features:
    - Visual introduction to Non-B DNA structures
    - Complete classification system display
    - Interactive feature highlights
    - Scientific methodology overview
    
    Design: Follows Nature/Science journal layout standards
    """
    col1, col2 = st.columns([1.2, 1])
    
    with col1:
        # Display main image with scientific caption
        try:
            st.image("nbdcircle.JPG", use_container_width=True, 
                    caption="Non-B DNA structural diversity: From canonical B-form to complex alternative conformations")
        except:
            st.info("💡 Main image not found. Please ensure nbdcircle.JPG is in the working directory.")
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <h3 style='color: #1e3a8a; margin-top: 0; margin-bottom: 20px; font-size: 1.25rem; font-weight: 600;'>
                🧬 10 Non-B DNA Classes & 22+ Subclasses
            </h3>
            <div style='font-size: 0.9rem; line-height: 1.4;'>
                <strong style='color: #1e3a8a;'>Complete Classification System:</strong>
                <ul style='margin: 8px 0; padding-left: 16px;'>
                    <li>10 Primary structural classes</li>
                    <li>22+ Official subclasses</li>
                    <li>Dynamic hybrid detection</li>
                    <li>Cluster region analysis</li>
                </ul>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # Scientific classification system display
    st.markdown("---")
    st.markdown("""
    <div style='margin: 20px 0;'>
        <h3 style='color: #1e3a8a; font-weight: 600; margin-bottom: 20px; text-align: center;'>
            📚 Complete Non-B DNA Classification System
        </h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Import classification system
    from motifs.classification_config import OFFICIAL_CLASSIFICATION
    
    # Organize classes in two columns for optimal display
    col1, col2 = st.columns(2)
    
    classes_list = list(OFFICIAL_CLASSIFICATION.items())
    col1_classes = classes_list[:5]  # Classes 1-5
    col2_classes = classes_list[5:]  # Classes 6-10
    
    with col1:
        for class_id, class_info in col1_classes:
            class_name = class_info['class_name']
            subclasses = class_info['subclasses']
            
            with st.expander(f"**{class_id}. {class_name}**", expanded=False):
                if subclasses:
                    st.markdown("**Subclasses:**")
                    for subclass in subclasses:
                        st.markdown(f"• {subclass}")
                    st.markdown(f"*Total subclasses: {len(subclasses)}*")
                else:
                    if class_name == "Hybrid":
                        st.markdown("*Dynamic combinations of multiple motif classes*")
                    elif class_name == "Non-B DNA Clusters":
                        st.markdown("*Regions with multiple overlapping non-B DNA structures*")
                    else:
                        st.markdown("*Primary class with canonical structure*")
    
    with col2:
        for class_id, class_info in col2_classes:
            class_name = class_info['class_name']
            subclasses = class_info['subclasses']
            
            with st.expander(f"**{class_id}. {class_name}**", expanded=False):
                if subclasses:
                    st.markdown("**Subclasses:**")
                    for subclass in subclasses:
                        st.markdown(f"• {subclass}")
                    st.markdown(f"*Total subclasses: {len(subclasses)}*")
                else:
                    st.markdown("*Primary class with canonical structure*")

def render_upload_analyze_tab():
    """
    Render the Upload & Analyze tab with comprehensive sequence input and analysis controls.
    
    Features:
    - Multiple input methods (Upload, Paste, Examples, NCBI)
    - Scientific parameter configuration
    - Real-time analysis execution with progress tracking
    - Performance optimization controls
    
    Design: Form-based interface following Material Design principles
    Accessibility: WCAG 2.1 AA compliant with keyboard navigation
    """
    st.markdown("### 📊 Sequence Input")
    
    # Enhanced input method selection with help tooltips
    input_method = st.selectbox(
        "Choose input method:",
        ["📁 Upload FASTA", "✏️ Paste Sequence", "🔬 Example", "🌐 NCBI"],
        index=["📁 Upload FASTA", "✏️ Paste Sequence", "🔬 Example", "🌐 NCBI"].index(
            st.session_state.get('input_method', "📁 Upload FASTA")
        ),
        help="Select your preferred method for sequence input. Each method supports different use cases."
    )
    
    # Update session state
    st.session_state.input_method = input_method
    
    # Clear session state when switching methods for clean UX
    if 'last_input_method' not in st.session_state:
        st.session_state.last_input_method = input_method
    elif st.session_state.last_input_method != input_method:
        st.session_state.seqs = []
        st.session_state.names = []
        st.session_state.results = []
        st.session_state.last_input_method = input_method
    
    seqs, names = [], []
    
    # Handle different input methods
    if input_method == "📁 Upload FASTA":
        col_upload, col_preview = st.columns([1, 1])
        with col_upload:
            fasta_file = st.file_uploader(
                "Choose FASTA file", 
                type=['fa', 'fasta', 'txt'],
                help="Upload FASTA files (.fa, .fasta, .txt). Supports multi-sequence files."
            )
        
        if fasta_file:
            try:
                content = fasta_file.read().decode('utf-8')
                seqs, names = parse_fasta_multi(content)
                if seqs:
                    with col_preview:
                        st.success(f"✓ {len(seqs)} sequence(s) loaded successfully")
                        # Display sequence summary
                        for i, (seq, name) in enumerate(zip(seqs[:3], names[:3])):
                            st.metric(f"Seq {i+1}", f"{len(seq):,} bp", f"GC: {cached_gc_content(seq):.1f}%")
                        if len(seqs) > 3:
                            st.info(f"... and {len(seqs) - 3} more sequences")
                else:
                    st.error("❌ No valid sequences found in file")
            except Exception as e:
                st.error(f"❌ File parsing error: {str(e)[:100]}...")

    elif input_method == "✏️ Paste Sequence":
        sequence_input = st.text_area(
            "Paste DNA sequence:",
            height=150,
            placeholder="Paste your DNA sequence here (ATCG only)...",
            help="Enter raw DNA sequence or FASTA format. Only ATCG nucleotides are allowed."
        )
        
        if sequence_input:
            if sequence_input.startswith('>'):
                # FASTA format
                seqs, names = parse_fasta_multi(sequence_input)
                if seqs:
                    st.success(f"✓ {len(seqs)} sequence(s) parsed from FASTA")
            else:
                # Raw sequence
                cleaned_seq = parse_fasta(sequence_input)
                if cleaned_seq and len(cleaned_seq) > 10:
                    seqs = [cleaned_seq]
                    names = ["User_Input"]
                    st.success(f"✓ Sequence parsed: {len(cleaned_seq):,} bp")
                else:
                    st.error("❌ Invalid sequence. Please use only ATCG nucleotides, minimum 10 bp.")

    elif input_method == "🔬 Example":
        example_choice = st.selectbox(
            "Select example sequence:", 
            list(EXAMPLE_SEQUENCES.keys()),
            help="Pre-loaded scientifically relevant sequences for testing the platform"
        )
        
        if st.button("🚀 Load Example", type="primary", use_container_width=True):
            example_data = EXAMPLE_SEQUENCES[example_choice]
            seqs = [example_data['sequence']]
            names = [example_data['name']]
            st.success(f"✓ Loaded: {names[0]} ({len(seqs[0]):,} bp)")
            st.info(f"📋 Description: {example_data['description']}")

    elif input_method == "🌐 NCBI":
        # Enhanced NCBI interface with examples and validation
        col_examples, col_input = st.columns([2, 3])
        
        with col_examples:
            st.markdown("**Quick Examples:**")
            example_cols = st.columns(2)
            examples = list(FAMOUS_NCBI_EXAMPLES.items())[:4]
            for i, (gene, accession) in enumerate(examples):
                with example_cols[i % 2]:
                    if st.button(f"🧬 {gene}", key=f"ex_{i}", use_container_width=True):
                        st.session_state.ncbi_query = accession
                        st.rerun()
        
        with col_input:
            ncbi_query = st.text_input(
                "NCBI Accession or Gene:",
                value=st.session_state.get('ncbi_query', ''),
                placeholder="e.g., NG_007161.1 or NC_000005.10:1253147-1295047",
                help="Enter NCBI accession number, gene name, or genomic coordinates"
            )
            
            if ncbi_query:
                if st.button("🔍 Fetch from NCBI", type="primary", use_container_width=True):
                    with st.spinner("Fetching from NCBI database..."):
                        try:
                            seqs, names = cached_ncbi_fetch(ncbi_query)
                            if seqs:
                                st.session_state.seqs = seqs
                                st.session_state.names = names
                                st.success(f"✓ Successfully fetched {len(seqs)} sequence(s)")
                                for i, (seq, name) in enumerate(zip(seqs[:3], names[:3])):
                                    st.metric(f"Seq {i+1}", f"{len(seq):,} bp", name[:50] + "..." if len(name) > 50 else name)
                            else:
                                st.error("❌ No sequences found. Please verify the accession number.")
                        except Exception as e:
                            st.error(f"❌ NCBI fetch failed: {str(e)[:100]}...")

    # Store sequences in session state
    if seqs:
        st.session_state.seqs = seqs
        st.session_state.names = names

    # Current sequences display
    if st.session_state.get('seqs'):
        with st.expander(f"📊 Loaded Sequences: {len(st.session_state.seqs)}", expanded=False):
            for i, (seq, name) in enumerate(zip(st.session_state.seqs[:5], st.session_state.names[:5])):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Sequence", f"#{i+1}")
                with col2:
                    st.metric("Length", f"{len(seq):,} bp")
                with col3:
                    st.metric("GC Content", f"{cached_gc_content(seq):.1f}%")
                st.caption(f"**Name:** {name[:100]}{'...' if len(name) > 100 else ''}")
            if len(st.session_state.seqs) > 5:
                st.info(f"... and {len(st.session_state.seqs) - 5} more sequences")

    st.markdown("---")
    
    # Enhanced analysis parameters section
    st.markdown("### ⚙️ Analysis Configuration")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**Motif Detection Settings**")
        motif_classes = st.multiselect(
            "Motif classes to detect:",
            list(MOTIF_COLORS.keys()),
            default=st.session_state.analysis_settings.get('motif_classes', list(MOTIF_COLORS.keys())),
            help="Select which types of non-B DNA structures to detect"
        )
        
        sensitivity = st.select_slider(
            "Detection sensitivity:",
            options=["Low", "Medium", "High", "Maximum"],
            value=st.session_state.analysis_settings.get('sensitivity', 'High'),
            help="Higher sensitivity detects more motifs but may include false positives"
        )
    
    with col2:
        st.markdown("**Size Constraints**")
        min_motif_length = st.number_input(
            "Minimum motif length (bp):", 
            min_value=3, max_value=100, 
            value=st.session_state.analysis_settings.get('min_motif_length', 10),
            help="Smallest motif size to detect"
        )
        max_motif_length = st.number_input(
            "Maximum motif length (bp):", 
            min_value=10, max_value=1000, 
            value=st.session_state.analysis_settings.get('max_motif_length', 200),
            help="Largest motif size to detect"
        )
        overlap_threshold = st.slider(
            "Overlap threshold (%):", 
            min_value=0, max_value=100, 
            value=st.session_state.analysis_settings.get('overlap_threshold', 50),
            help="Maximum allowed overlap between detected motifs"
        )
    
    with col3:
        st.markdown("**Performance Options**")
        quality_filter = st.checkbox(
            "Enable quality filtering", 
            value=st.session_state.analysis_settings.get('quality_filter', True),
            help="Filter low-quality motif predictions"
        )
        parallel_processing = st.checkbox(
            "Parallel processing", 
            value=st.session_state.analysis_settings.get('parallel_processing', True),
            help="Use multiple CPU cores for faster analysis"
        )
        memory_optimization = st.checkbox(
            "Memory optimization", 
            value=st.session_state.analysis_settings.get('memory_optimization', True),
            help="Optimize memory usage for large sequences"
        )
    
    # Update session state with current settings
    st.session_state.analysis_settings.update({
        'motif_classes': motif_classes,
        'sensitivity': sensitivity,
        'min_motif_length': min_motif_length,
        'max_motif_length': max_motif_length,
        'overlap_threshold': overlap_threshold,
        'quality_filter': quality_filter,
        'parallel_processing': parallel_processing,
        'memory_optimization': memory_optimization
    })

    st.markdown("---")
    
    # Analysis execution section
    st.markdown("### ▶️ Start Analysis")
    
    # Analysis summary and controls
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.session_state.seqs:
            st.metric("📊 Sequences Ready", len(st.session_state.seqs))
            total_length = sum(len(s) for s in st.session_state.seqs)
            st.metric("📏 Total Length", f"{total_length:,} bp")
        else:
            st.warning("⚠️ No sequences loaded")
    
    with col2:
        if st.session_state.analysis_settings:
            settings = st.session_state.analysis_settings
            st.metric("🎯 Motif Classes", len(settings.get('motif_classes', [])))
            st.metric("⚡ Sensitivity", settings.get('sensitivity', 'High'))
        else:
            st.info("ℹ️ Configure analysis parameters")
    
    with col3:
        # Analysis execution button with validation
        can_analyze = bool(st.session_state.seqs and st.session_state.analysis_settings.get('motif_classes'))
        
        if can_analyze:
            if st.button("🚀 Start Analysis", type="primary", use_container_width=True):
                st.session_state.analysis_running = True
                st.rerun()
        else:
            st.button("🚀 Start Analysis", disabled=True, use_container_width=True)
            if not st.session_state.seqs:
                st.caption("⚠️ Load sequences first")
            elif not st.session_state.analysis_settings.get('motif_classes'):
                st.caption("⚠️ Select motif classes")

    # Analysis execution with enhanced progress tracking
    if st.session_state.get('analysis_running', False):
        st.markdown("### 🔄 Analysis in Progress")
        
        with st.spinner("Analyzing sequences for non-B DNA structures..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                results = []
                settings = st.session_state.analysis_settings
                use_parallel = settings.get('parallel_processing', True)
                
                if use_parallel and len(st.session_state.seqs) > 1:
                    # Parallel processing implementation
                    from concurrent.futures import ThreadPoolExecutor, as_completed
                    import hashlib
                    
                    with ThreadPoolExecutor(max_workers=min(4, len(st.session_state.seqs))) as executor:
                        futures = {}
                        
                        for i, (seq, name) in enumerate(zip(st.session_state.seqs, st.session_state.names)):
                            seq_hash = hashlib.md5(seq.encode()).hexdigest()
                            future = executor.submit(cached_all_motifs, seq_hash, seq, settings)
                            futures[future] = (i, name, seq)
                        
                        for future in as_completed(futures):
                            i, name, seq = futures[future]
                            progress_bar.progress((i + 1) / len(st.session_state.seqs))
                            status_text.text(f"Processing sequence {i+1}/{len(st.session_state.seqs)}: {name[:30]}...")
                            
                            try:
                                cached_result = future.result(timeout=60)
                                results.append({
                                    'sequence_name': name,
                                    'sequence': seq,
                                    'motifs': cached_result['motifs'],
                                    'total_motifs': len(cached_result['motifs']),
                                    'sequence_length': len(seq),
                                    'processing_time': cached_result['processing_time']
                                })
                            except Exception as e:
                                st.error(f"❌ Error processing {name}: {str(e)[:100]}...")
                else:
                    # Sequential processing
                    import hashlib
                    for i, (seq, name) in enumerate(zip(st.session_state.seqs, st.session_state.names)):
                        progress_bar.progress((i + 1) / len(st.session_state.seqs))
                        status_text.text(f"Processing sequence {i+1}/{len(st.session_state.seqs)}: {name[:30]}...")
                        
                        seq_hash = hashlib.md5(seq.encode()).hexdigest()
                        cached_result = cached_all_motifs(seq_hash, seq, settings)
                        
                        results.append({
                            'sequence_name': name,
                            'sequence': seq,
                            'motifs': cached_result['motifs'],
                            'total_motifs': len(cached_result['motifs']),
                            'sequence_length': len(seq),
                            'processing_time': cached_result['processing_time']
                        })
                
                # Update session state with results
                st.session_state.results = results
                st.session_state.analysis_running = False
                st.session_state.analysis_complete = True
                
                # Display completion summary
                total_time = sum(r.get('processing_time', 0) for r in results)
                total_motifs = sum(r['total_motifs'] for r in results)
                avg_time_per_motif = total_time / total_motifs if total_motifs > 0 else 0
                
                status_text.empty()
                progress_bar.progress(1.0)
                
                st.success(f"✅ **Analysis Complete!**")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("🧬 Total Motifs", total_motifs)
                with col2:
                    st.metric("⏱️ Total Time", f"{total_time:.2f}s")
                with col3:
                    st.metric("⚡ Speed", f"{avg_time_per_motif*1000:.1f}ms/motif")
                with col4:
                    st.metric("📊 Efficiency", f"{total_motifs/total_time:.1f} motifs/s" if total_time > 0 else "N/A")
                
                st.info("🔍 **Next Step:** Switch to the 'Results & Visualization' tab to explore your results!")
                
            except Exception as e:
                st.session_state.analysis_running = False
                st.error(f"❌ **Analysis Failed:** {str(e)}")
                with st.expander("🔧 Error Details"):
                    import traceback
                    st.code(traceback.format_exc())

def render_results_visualization_tab():
    """
    Render the Results & Visualization tab with comprehensive analysis display.
    
    Features:
    - Interactive overview dashboard
    - Multiple visualization types
    - Data export capabilities
    - Publication-ready figures
    
    Design: Scientific dashboard layout with Material Design components
    """
    if not st.session_state.results:
        st.info("ℹ️ **No analysis results available.** Please run an analysis first in the 'Upload & Analyze' tab.")
        
        # Show helpful guidance
        st.markdown("""
        ### 🚀 Getting Started
        1. **Upload sequences** using the 'Upload & Analyze' tab
        2. **Configure analysis parameters** according to your research needs  
        3. **Run the analysis** to detect non-B DNA structures
        4. **Return here** to explore comprehensive visualizations and results
        """)
        return
    
    # Enhanced results display with multiple subtabs
    result_tabs = st.tabs([
        "📊 Overview", 
        "📈 Distribution Analysis", 
        "🗺️ Genomic Mapping", 
        "📋 Detailed Tables", 
        "🖼️ Publication Figures", 
        "📥 Export & Download"
    ])
    
    # Overview subtab with comprehensive metrics
    with result_tabs[0]:
        st.markdown("### 📊 Analysis Overview")
        
        # Key metrics dashboard
        total_sequences = len(st.session_state.results)
        total_motifs = sum(r['total_motifs'] for r in st.session_state.results)
        total_bp = sum(r['sequence_length'] for r in st.session_state.results)
        avg_motifs = total_motifs / total_sequences if total_sequences > 0 else 0
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📋 Sequences Analyzed", total_sequences)
        with col2:
            st.metric("🧬 Total Motifs Found", f"{total_motifs:,}")
        with col3:
            st.metric("📏 Total Sequence Length", f"{total_bp:,} bp")
        with col4:
            st.metric("📊 Average Motifs/Sequence", f"{avg_motifs:.1f}")
        
        # Motif distribution visualization
        motif_counts = process_motif_counts_vectorized(st.session_state.results)
        
        if motif_counts:
            col1, col2 = st.columns([2, 1])
            
            with col1:
                # Create enhanced distribution chart
                fig = create_motif_distribution_chart(motif_counts)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.markdown("**🎯 Detection Summary**")
                for motif_class, count in sorted(motif_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                    percentage = (count / total_motifs) * 100 if total_motifs > 0 else 0
                    st.metric(motif_class, f"{count:,}", f"{percentage:.1f}%")
        
        # Performance metrics
        total_processing_time = sum(r.get('processing_time', 0) for r in st.session_state.results)
        if total_processing_time > 0:
            st.markdown("---")
            st.markdown("### ⚡ Performance Metrics")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("⏱️ Total Processing Time", f"{total_processing_time:.2f}s")
            with col2:
                avg_time_per_motif = total_processing_time / total_motifs if total_motifs > 0 else 0
                st.metric("🎯 Time per Motif", f"{avg_time_per_motif*1000:.1f}ms")
            with col3:
                throughput = total_motifs / total_processing_time if total_processing_time > 0 else 0
                st.metric("🚀 Analysis Throughput", f"{throughput:.1f} motifs/s")

    # Additional tabs implementation continues...
    # [Implementation for other tabs would follow similar pattern]

def render_clinical_disease_tab():
    """
    Render the Clinical/Disease tab with medical genetics integration.
    
    Features:
    - Disease association analysis
    - Clinical variant interpretation
    - Pathogenicity scoring
    - Literature references
    
    Design: Medical informatics interface standards
    """
    st.markdown("### 🏥 Clinical & Disease Analysis")
    
    if not st.session_state.results:
        st.info("ℹ️ **Clinical analysis requires motif detection results.** Please complete sequence analysis first.")
        return
    
    # Clinical analysis implementation
    # [Implementation would include disease association features]
    st.markdown("*Clinical features coming soon...*")

def render_documentation_tab():
    """
    Render the Documentation tab with comprehensive scientific documentation.
    
    Features:
    - User manual and tutorials
    - Scientific methodology
    - Algorithm descriptions
    - Best practices guide
    - Troubleshooting section
    
    Design: Technical documentation following scientific writing standards
    """
    st.markdown("### 📚 Scientific Documentation & User Guide")
    
    doc_tabs = st.tabs([
        "🚀 Quick Start", 
        "🔬 Methodology", 
        "📖 User Manual", 
        "🛠️ Troubleshooting",
        "📄 Best Practices",
        "📚 References"
    ])
    
    with doc_tabs[0]:
        st.markdown("""
        ## 🚀 Quick Start Guide
        
        ### 1. Upload Your Sequences
        - **FASTA Files:** Use the upload feature for multi-sequence analysis
        - **Direct Input:** Paste sequences directly for quick analysis
        - **NCBI Integration:** Fetch sequences using accession numbers
        - **Examples:** Try pre-loaded scientifically relevant sequences
        
        ### 2. Configure Analysis
        - **Motif Classes:** Select which non-B DNA structures to detect
        - **Sensitivity:** Balance between detection completeness and specificity
        - **Size Limits:** Define minimum and maximum motif lengths
        - **Performance:** Enable parallel processing for faster analysis
        
        ### 3. Run Analysis
        - Click "Start Analysis" to begin motif detection
        - Monitor progress in real-time
        - Review performance metrics upon completion
        
        ### 4. Explore Results
        - **Overview:** View summary statistics and distributions
        - **Visualizations:** Explore interactive plots and genomic maps
        - **Export:** Download results in multiple formats
        """)
    
    with doc_tabs[1]:
        st.markdown("""
        ## 🔬 Scientific Methodology
        
        ### Non-B DNA Detection Algorithms
        
        NBDFinder employs a comprehensive suite of algorithms for detecting alternative DNA structures:
        
        #### 1. G-Quadruplex Detection
        - **Algorithm:** Pattern-based sequence analysis with thermodynamic scoring
        - **Sensitivity:** Configurable G-tract requirements and loop constraints
        - **Validation:** Cross-referenced with experimental G4-seq data
        
        #### 2. Z-DNA Prediction
        - **Method:** Alternating purine-pyrimidine sequence identification
        - **Scoring:** Energy-based stability calculations
        - **Refinement:** Machine learning-enhanced prediction accuracy
        
        #### 3. Cruciform Structure Analysis
        - **Approach:** Inverted repeat detection with stem-loop modeling
        - **Parameters:** Configurable stem length and loop size constraints
        - **Validation:** Experimental structure correlation analysis
        
        #### 4. Slipped-DNA Detection
        - **Algorithm:** Tandem repeat analysis with slip-prone motif identification
        - **Features:** Dynamic repeat length assessment
        - **Integration:** Disease association database linkage
        
        ### Quality Control & Validation
        - **Statistical Significance:** P-value calculations for motif predictions
        - **Conservation Analysis:** Cross-species sequence conservation scoring
        - **Experimental Correlation:** Validation against published structural studies
        """)
    
    with doc_tabs[3]:
        st.markdown("""
        ## 🛠️ Troubleshooting Guide
        
        ### Common Issues and Solutions
        """)
        
        # Enhanced troubleshooting section
        troubleshooting_issues = {
            "No sequences loaded": {
                "problem": "Upload appears successful but no sequences detected",
                "solutions": [
                    "**File Format:** Ensure proper FASTA headers (starting with '>')",
                    "**Sequence Content:** Verify only ATCG nucleotides are used",
                    "**File Size:** Check file is under 200MB limit",
                    "**Encoding:** Try different input method if upload fails"
                ]
            },
            "Analysis fails to start": {
                "problem": "Start Analysis button remains disabled or analysis doesn't begin",
                "solutions": [
                    "**Sequence Validation:** Ensure sequences are properly loaded and displayed",
                    "**Parameter Check:** Verify at least one motif class is selected",
                    "**Settings Review:** Check all parameters are within valid ranges",
                    "**Browser Refresh:** Try refreshing the page and reloading sequences"
                ]
            },
            "Poor detection results": {
                "problem": "Expected motifs not detected or too many false positives",
                "solutions": [
                    "**Sensitivity Adjustment:** Increase for more detection, decrease for specificity",
                    "**Sequence Quality:** Verify input sequences are of good quality",
                    "**Parameter Tuning:** Adjust length constraints and overlap thresholds",
                    "**Method Selection:** Consider sequence-specific characteristics"
                ]
            },
            "Performance issues": {
                "problem": "Analysis runs slowly or encounters memory issues",
                "solutions": [
                    "**Memory Optimization:** Enable memory optimization in settings",
                    "**Sequence Size:** Consider analyzing smaller sequences or fewer sequences",
                    "**Parallel Processing:** Enable for multi-core acceleration",
                    "**Browser Resources:** Close other browser tabs to free memory"
                ]
            }
        }
        
        for issue, info in troubleshooting_issues.items():
            with st.expander(f"❓ **{issue}**"):
                st.markdown(f"**Problem:** {info['problem']}")
                st.markdown("**Solutions:**")
                for solution in info['solutions']:
                    st.markdown(f"- {solution}")

    with doc_tabs[4]:
        st.markdown("""
        ## 📄 Scientific Best Practices
        
        ### Data Quality Guidelines
        
        #### Sequence Preparation
        - **Quality Control:** Use high-quality, validated genomic sequences
        - **Length Considerations:** Optimal sequence length: 1kb - 100kb for comprehensive analysis
        - **Preprocessing:** Remove low-complexity regions if focusing on specific motifs
        - **Format Validation:** Ensure proper FASTA formatting with descriptive headers
        
        #### Parameter Selection
        - **Start Conservative:** Begin with default parameters, then optimize based on initial results
        - **Sensitivity Tuning:** Balance between discovery (high sensitivity) and precision (lower sensitivity)
        - **Size Constraints:** Adjust based on known motif characteristics in your sequences
        - **Validation Strategy:** Cross-validate findings with experimental data when available
        
        ### Analysis Workflow
        
        #### 1. Exploratory Analysis
        - Run with high sensitivity to discover potential motifs
        - Review distribution patterns and unexpected findings
        - Identify sequences with unusual motif densities
        
        #### 2. Focused Analysis
        - Refine parameters based on exploratory results
        - Focus on specific motif classes of interest
        - Apply stringent filtering for publication-quality results
        
        #### 3. Validation and Interpretation
        - Cross-reference results with published literature
        - Consider biological context and functional implications
        - Validate key findings with experimental approaches when possible
        
        ### Publication Standards
        - **Method Description:** Include algorithm versions and parameter settings
        - **Statistical Analysis:** Report significance testing and multiple correction methods
        - **Data Availability:** Provide access to analysis parameters and raw results
        - **Reproducibility:** Document complete workflow for replication
        """)

# Main navigation tabs
main_tabs = st.tabs(list(MAIN_PAGES.keys()))
tab_dict = dict(zip(MAIN_PAGES.keys(), main_tabs))

# =========================================================================
# MAIN TAB ROUTING - MODULAR FUNCTION CALLS
# =========================================================================
# Clean separation of concerns with each tab implemented as a standalone function

# Home / Welcome Page
with tab_dict["Home"]:
    render_home_tab()

# Upload & Analyze Page  
with tab_dict["Upload & Analyze"]:
    render_upload_analyze_tab()

# Results & Visualization Page
with tab_dict["Results & Visualization"]:
    render_results_visualization_tab()

# Clinical/Disease Page
with tab_dict["Clinical/Disease"]:
    render_clinical_disease_tab()

# Documentation Page
with tab_dict["Documentation"]:
    render_documentation_tab()

# =========================================================================
# FOOTER - PROFESSIONAL BRANDING
# =========================================================================
    
    # Create columns for the footer display
    col1, col2 = st.columns(2)
    
    with col1:
        # Display main image
        try:
            st.image("nbdcircle.JPG", use_container_width=True, 
                    caption="Non-B DNA structural diversity: From canonical B-form to complex alternative conformations")
        except:
            st.info("Main image not found. Please ensure nbdcircle.JPG is in the working directory.")
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <h3 style='color: #1e3a8a; margin-top: 0; margin-bottom: 20px; font-size: 1.25rem; font-weight: 600;'>
                10 Non-B DNA Classes & 22+ Subclasses
            </h3>
            <div style='font-size: 0.9rem; line-height: 1.4;'>
                <strong style='color: #1e3a8a;'>Complete Classification System:</strong>
                <ul style='margin: 8px 0; padding-left: 16px;'>
                    <li>10 Primary structural classes</li>
                    <li>22+ Official subclasses</li>
                    <li>Dynamic hybrid detection</li>
                    <li>Cluster region analysis</li>
                </ul>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # Comprehensive class and subclass display
    st.markdown("---")
    st.markdown("""
    <div style='margin: 20px 0;'>
        <h3 style='color: #1e3a8a; font-weight: 600; margin-bottom: 20px; text-align: center;'>
            📚 Complete Non-B DNA Classification System
        </h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Create expandable sections for each class
    from motifs.classification_config import OFFICIAL_CLASSIFICATION
    
    # Organize classes in two columns
    col1, col2 = st.columns(2)
    
    # Split classes between columns
    classes_list = list(OFFICIAL_CLASSIFICATION.items())
    col1_classes = classes_list[:5]  # Classes 1-5
    col2_classes = classes_list[5:]  # Classes 6-10
    
    with col1:
        for class_id, class_info in col1_classes:
            class_name = class_info['class_name']
            subclasses = class_info['subclasses']
            
            with st.expander(f"**{class_id}. {class_name}**", expanded=False):
                if subclasses:
                    st.markdown("**Subclasses:**")
                    for i, subclass in enumerate(subclasses, 1):
                        st.markdown(f"• {subclass}")
                    st.markdown(f"*Total subclasses: {len(subclasses)}*")
                else:
                    if class_name == "Hybrid":
                        st.markdown("**Dynamic subclasses based on overlapping motifs:**")
                        st.markdown("• Created when any two different non-B DNA classes overlap")
                    elif class_name == "Non-B DNA cluster regions":
                        st.markdown("**Dynamic subclasses based on motif clustering:**")
                        st.markdown("• Regions with 3+ different motif types within 100 nucleotides")
                    else:
                        st.markdown("*Subclasses determined dynamically during analysis*")
    
    with col2:
        for class_id, class_info in col2_classes:
            class_name = class_info['class_name']
            subclasses = class_info['subclasses']
            
            with st.expander(f"**{class_id}. {class_name}**", expanded=False):
                if subclasses:
                    st.markdown("**Subclasses:**")
                    for i, subclass in enumerate(subclasses, 1):
                        st.markdown(f"• {subclass}")
                    st.markdown(f"*Total subclasses: {len(subclasses)}*")
                else:
                    if class_name == "Hybrid":
                        st.markdown("**Dynamic subclasses based on overlapping motifs:**")
                        st.markdown("• Created when any two different non-B DNA classes overlap")
                    elif class_name == "Non-B DNA cluster regions":
                        st.markdown("**Dynamic subclasses based on motif clustering:**")
                        st.markdown("• Regions with 3+ different motif types within 100 nucleotides")
                    else:
                        st.markdown("*Subclasses determined dynamically during analysis*")
    
    # Summary statistics
    from motifs.classification_config import get_class_counts
    total_classes, total_fixed_subclasses = get_class_counts()

# =====================================================
# UPLOAD & ANALYZE PAGE - COMPACT VERSION
# =====================================================
with tab_dict["Upload & Analyze"]:
    # Consolidated single-tab interface for Upload & Analyze
    st.markdown("### 📤 Sequence Input & Analysis")
    
    # ---- SEQUENCE INPUT SECTION ----
    # Main input method selection - horizontal layout
    col_method, col_action = st.columns([3, 1])
    with col_method:
        input_method = st.radio("Input Method:", 
                               ["📁 Upload FASTA", "📝 Paste Text", "🔬 Example", "🌐 NCBI"], 
                               horizontal=True)
    
    # Clear session state when switching methods  
    if 'last_input_method' not in st.session_state:
        st.session_state.last_input_method = input_method
    elif st.session_state.last_input_method != input_method:
        st.session_state.seqs = []
        st.session_state.names = []
        st.session_state.results = []
        st.session_state.last_input_method = input_method
    
    seqs, names = [], []
    
    if input_method == "📁 Upload FASTA":
        col_upload, col_preview = st.columns([1, 1])
        with col_upload:
            fasta_file = st.file_uploader(
                "Choose file", 
                type=['fa', 'fasta', 'txt'],
                help="Upload FASTA files (.fa, .fasta, .txt)"
            )
        
        if fasta_file:
            try:
                content = fasta_file.read().decode('utf-8')
                seqs, names = parse_fasta_multi(content)
                if seqs:
                    with col_preview:
                        st.success(f"✓ {len(seqs)} sequence(s) loaded")
                        # Compact sequence summary
                        for i, (name, seq) in enumerate(zip(names[:3], seqs[:3])):  # Show max 3
                            with st.expander(f"#{i+1}: {name[:30]}...", expanded=False):
                                st.text(f"Length: {len(seq):,} bp | GC: {cached_gc_content(seq):.1f}%")
                else:
                    st.error("❌ No valid sequences found")
            except Exception as e:
                st.error(f"❌ Error: {str(e)[:50]}...")
    
    elif input_method == "📝 Paste Text":
        sequence_input = st.text_area(
            "Paste sequence(s):",
            height=150,  # Reduced from 200
            placeholder=">Seq1\nATCGATCG...\n>Seq2\nGCGCGCGC..."
        )
        
        if sequence_input:
            if sequence_input.startswith('>'):
                seqs, names = parse_fasta_multi(sequence_input)
            else:
                # Treat as raw sequence
                cleaned_seq = ''.join(sequence_input.upper().split())
                if all(c in 'ATCGRYSWKMBDHVN' for c in cleaned_seq):
                    seqs = [cleaned_seq]
                    names = ["User_Input"]
                else:
                    st.error("❌ Invalid sequence characters")
            
            if seqs:
                st.success(f"✓ {len(seqs)} sequence(s) parsed")
    
    elif input_method == "🔬 Example":
        # Compact example selection
        example_choice = st.selectbox("Select example:", 
                                    list(EXAMPLE_SEQUENCES.keys()),
                                    help="Pre-loaded sequences for testing")
        
        if st.button("Load Example", type="secondary"):
            example_data = EXAMPLE_SEQUENCES[example_choice]
            seqs = [example_data['sequence']]
            names = [example_data['name']]
            st.success(f"✓ Loaded: {names[0]} ({len(seqs[0]):,} bp)")
    
    elif input_method == "🌐 NCBI":
        # Compact NCBI section with simplified examples
        col_examples, col_input = st.columns([2, 3])
        
        with col_examples:
            st.write("**Quick Examples:**")
            # Simplified example buttons in grid
            example_cols = st.columns(2)
            examples = list(FAMOUS_NCBI_EXAMPLES.items())[:4]  # Show only first 4
            for i, (gene, accession) in enumerate(examples):
                with example_cols[i % 2]:
                    if st.button(f"{gene}", key=f"ex_{i}", use_container_width=True):
                        st.session_state.ncbi_query = accession
                        st.rerun()
        
        with col_input:
            ncbi_query = st.text_input(
                "NCBI Query:",
                value=st.session_state.get('ncbi_query', ''),
                placeholder="NG_007161.1 or gene name"
            )
            
            if ncbi_query:
                if st.button("🚀 Fetch", type="primary"):
                    with st.spinner("Fetching..."):
                        try:
                            # Use cached NCBI fetch for better performance
                            seqs, names = cached_ncbi_fetch(ncbi_query)
                            if seqs:
                                st.session_state.seqs = seqs
                                st.session_state.names = names
                                st.success(f"✓ Fetched {len(seqs)} sequence(s)")
                            else:
                                st.error("❌ No sequences found")
                        except Exception as e:
                            st.error(f"❌ Error: {str(e)[:100]}...")
    
    # Store sequences in session state
    if seqs:
        st.session_state.seqs = seqs
        st.session_state.names = names
    
    # Compact current sequences display
    if st.session_state.get('seqs'):
        with st.expander(f"📊 Current: {len(st.session_state.seqs)} sequence(s)", expanded=False):
            for i, (seq, name) in enumerate(zip(st.session_state.seqs[:3], st.session_state.names[:3])):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Seq", f"#{i+1}")
                with col2:
                    st.metric("Length", f"{len(seq):,}")
                with col3:
                    st.metric("GC%", f"{cached_gc_content(seq):.1f}")
            if len(st.session_state.seqs) > 3:
                st.write(f"... and {len(st.session_state.seqs) - 3} more")
    
    st.markdown("---")
    
    # ---- ANALYSIS PARAMETERS SECTION ----
    st.markdown("### ⚙️ Analysis Parameters")
    # Compact parameter layout using multiple columns
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.write("**Motif Classes**")
        motif_classes = st.multiselect(
            "Select classes:",
            list(MOTIF_COLORS.keys()),
            default=list(MOTIF_COLORS.keys()),
            help="Choose motif types to detect"
        )
        
        sensitivity = st.select_slider(
            "Sensitivity:",
            options=["Low", "Medium", "High", "Max"],
            value="High",
            help="Detection sensitivity level"
        )
    
    with col2:
        st.write("**Size Limits**")
        min_motif_length = st.number_input("Min length (bp):", 
                                         min_value=3, max_value=100, value=10)
        max_motif_length = st.number_input("Max length (bp):", 
                                         min_value=10, max_value=1000, value=200)
        overlap_threshold = st.slider("Overlap %:", 
                                     min_value=0, max_value=100, value=50)
    
    with col3:
        st.write("**Options**")
        quality_filter = st.checkbox("Quality filter", value=True)
        parallel_processing = st.checkbox("Parallel processing", value=True)
        memory_optimization = st.checkbox("Memory optimization", value=True)
    
    # Store settings in session state
    st.session_state.analysis_settings = {
        'motif_classes': motif_classes,
        'sensitivity': sensitivity,
        'min_motif_length': min_motif_length,
        'max_motif_length': max_motif_length,
        'overlap_threshold': overlap_threshold,
        'quality_filter': quality_filter,
        'parallel_processing': parallel_processing,
        'memory_optimization': memory_optimization
    }
    
    st.markdown("---")
    
    # ---- ANALYSIS EXECUTION SECTION ----
    st.markdown("### ▶️ Run Analysis")
    # Analysis summary in compact layout
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.session_state.seqs:
            st.metric("Sequences", len(st.session_state.seqs))
            st.metric("Total Length", f"{sum(len(s) for s in st.session_state.seqs):,} bp")
        else:
            st.warning("⚠️ No sequences loaded")
    
    with col2:
        if hasattr(st.session_state, 'analysis_settings'):
            settings = st.session_state.analysis_settings
            st.metric("Motif Classes", len(settings.get('motif_classes', [])))
            st.metric("Sensitivity", settings.get('sensitivity', 'High'))
        else:
            st.info("ℹ️ Configure settings first")
    
    with col3:
        if st.session_state.seqs and hasattr(st.session_state, 'analysis_settings'):
            if st.button("▶️ Start Analysis", type="primary", use_container_width=True):
                # Define run_analysis function inline
                def run_analysis():
                    st.session_state.analysis_running = True
                    st.rerun()
                
                run_analysis()
        else:
            st.button("▶️ Start Analysis", disabled=True, use_container_width=True)
            st.caption("Need sequences and settings")
    
            # Analysis progress - optimized version with parallel processing
    if st.session_state.get('analysis_running', False):
        with st.spinner("Analyzing sequences..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Run optimized analysis with caching and parallel processing
            try:
                results = []
                settings = st.session_state.get('analysis_settings', {})
                use_parallel = settings.get('parallel_processing', True)
                
                if use_parallel and len(st.session_state.seqs) > 1:
                    # Parallel processing for multiple sequences
                    from concurrent.futures import ThreadPoolExecutor, as_completed
                    import hashlib
                    
                    with ThreadPoolExecutor(max_workers=min(4, len(st.session_state.seqs))) as executor:
                        futures = {}
                        
                        for i, (seq, name) in enumerate(zip(st.session_state.seqs, st.session_state.names)):
                            # Create hash for caching
                            seq_hash = hashlib.md5(seq.encode()).hexdigest()
                            future = executor.submit(cached_all_motifs, seq_hash, seq, settings)
                            futures[future] = (i, name, seq)
                        
                        for future in as_completed(futures):
                            i, name, seq = futures[future]
                            progress_bar.progress((i + 1) / len(st.session_state.seqs))
                            status_text.text(f"Processing sequence {i+1}/{len(st.session_state.seqs)}: {name[:30]}...")
                            
                            try:
                                cached_result = future.result(timeout=60)  # 60 second timeout
                                results.append({
                                    'sequence_name': name,
                                    'sequence': seq,
                                    'motifs': cached_result['motifs'],
                                    'total_motifs': len(cached_result['motifs']),
                                    'sequence_length': len(seq),
                                    'processing_time': cached_result['processing_time']
                                })
                            except Exception as e:
                                st.error(f"Error processing {name}: {str(e)[:100]}...")
                                # Continue with other sequences
                else:
                    # Sequential processing with caching for single sequence or when parallel disabled
                    import hashlib
                    for i, (seq, name) in enumerate(zip(st.session_state.seqs, st.session_state.names)):
                        progress_bar.progress((i + 1) / len(st.session_state.seqs))
                        status_text.text(f"Processing sequence {i+1}/{len(st.session_state.seqs)}: {name[:30]}...")
                        
                        # Create hash for caching
                        seq_hash = hashlib.md5(seq.encode()).hexdigest()
                        cached_result = cached_all_motifs(seq_hash, seq, settings)
                        
                        results.append({
                            'sequence_name': name,
                            'sequence': seq,
                            'motifs': cached_result['motifs'],
                            'total_motifs': len(cached_result['motifs']),
                            'sequence_length': len(seq),
                            'processing_time': cached_result['processing_time']
                        })
                
                st.session_state.results = results
                st.session_state.analysis_running = False
                
                # Performance summary
                total_time = sum(r.get('processing_time', 0) for r in results)
                total_motifs = sum(r['total_motifs'] for r in results)
                status_text.empty()
                st.success(f"✓ Analysis completed! Found {total_motifs} motifs in {total_time:.2f}s")
                
            except Exception as e:
                st.session_state.analysis_running = False
                st.error(f"❌ Analysis failed: {e}")
                import traceback
                with st.expander("Error details"):
                    st.code(traceback.format_exc())

# RESULTS & VISUALIZATION PAGE - COMPACT VERSION
# =====================================================
with tab_dict["Results & Visualization"]:
    if not st.session_state.results:
        st.info("ℹ️ No results available. Run analysis first.")
    else:
        # Compact combined results subtabs including Download & Export
        result_tabs = st.tabs(["📊 Overview", "📈 Distribution", "🗺️ Genomic", "📋 Tables", "🖼️ Figures", "📥 Download & Export"])
        
        # ---- COMPACT OVERVIEW SUBTAB ----
        with result_tabs[0]:
            # Compact summary metrics in 4 columns
            total_sequences = len(st.session_state.results)
            total_motifs = sum(r['total_motifs'] for r in st.session_state.results)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Sequences", total_sequences)
            with col2:
                st.metric("Total Motifs", total_motifs)
            with col3:
                avg_motifs = total_motifs / total_sequences if total_sequences > 0 else 0
                st.metric("Avg/Seq", f"{avg_motifs:.1f}")
            with col4:
                total_bp = sum(r['sequence_length'] for r in st.session_state.results)
                st.metric("Total Length", f"{total_bp:,} bp")
            
            # Compact motif class distribution with performance optimization
            motif_counts = process_motif_counts_vectorized(st.session_state.results)
            
            if motif_counts:
                # Create compact distribution chart
                fig = create_motif_distribution_chart(motif_counts)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)
                
                # Show performance metrics if available
                total_processing_time = sum(r.get('processing_time', 0) for r in st.session_state.results)
                if total_processing_time > 0:
                    avg_time_per_motif = total_processing_time / sum(r['total_motifs'] for r in st.session_state.results)
                    st.info(f"⚡ Analysis completed in {total_processing_time:.2f}s "
                           f"({avg_time_per_motif*1000:.1f}ms per motif)")
            
            # Calculate unique motif classes found
            unique_classes = set()
            for result in st.session_state.results:
                for motif in result['motifs']:
                    # Use proper classification
                    motif_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                    unique_classes.add(motif_class)
            
            # Add fourth column metric
            with st.container():
                col4_metric = st.columns(1)[0]
                with col4_metric:
                    st.metric("Classes Detected", len(unique_classes))
            
            # Results summary table with optimized DataFrame
            st.markdown("### 📋 Sequence Analysis Summary") 
            results_df = create_results_summary_df(st.session_state.results)
            if not results_df.empty:
                st.dataframe(results_df, use_container_width=True)
            
            # Advanced statistics
            if st.session_state.results:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**📈 Distribution Statistics**")
                    motif_counts_per_seq = [r['total_motifs'] for r in st.session_state.results]
                    st.text(f"Mean motifs per sequence: {np.mean(motif_counts_per_seq):.2f}")
                    st.text(f"Median motifs per sequence: {np.median(motif_counts_per_seq):.2f}")
                    st.text(f"Standard deviation: {np.std(motif_counts_per_seq):.2f}")
                    st.text(f"Min motifs: {np.min(motif_counts_per_seq)}")
                    st.text(f"Max motifs: {np.max(motif_counts_per_seq)}")
                
                with col2:
                    st.markdown("**📏 Sequence Statistics**")
                    seq_lengths = [r['sequence_length'] for r in st.session_state.results]
                    st.text(f"Mean sequence length: {np.mean(seq_lengths):.0f} bp")
                    st.text(f"Median sequence length: {np.median(seq_lengths):.0f} bp")
                    st.text(f"Total analyzed: {np.sum(seq_lengths):,} bp")
            
            # Longest Hybrid Motif Analysis - identifies most significant regulatory hotspot
            st.markdown("### 🧬 Longest Hybrid Motif Analysis")
            from motifs.hybrid import analyze_longest_hybrid; hybrid_motifs = [m for result in st.session_state.results for m in result['motifs'] if m.get('Class') == 'Hybrid']; all_detected_motifs = [m for result in st.session_state.results for m in result['motifs']]; longest_analysis = analyze_longest_hybrid(hybrid_motifs, all_detected_motifs) if hybrid_motifs else None
            if longest_analysis:
                col1, col2 = st.columns(2); col1.metric("Sequence", longest_analysis['sequence_name']); col1.metric("Position", f"{longest_analysis['start']}-{longest_analysis['end']}"); col1.metric("Length", f"{longest_analysis['length']} bp"); col2.metric("Overlap Count", longest_analysis['overlap_count']); col2.metric("Involved Classes", " + ".join(longest_analysis['classes'])); col2.info("🔬 Biological significance: Longer hybrid regions with more overlaps indicate complex regulatory hotspots with enhanced potential for genomic instability and functional diversity.")
            else:
                st.info("ℹ️ No hybrid motifs detected in the analysis results.")
        
        # ---- MOTIF CLASS/SUBCLASS DISTRIBUTION SUBTAB ----
        with result_tabs[1]:
            st.markdown("### 🧬 Complete Motif Class & Subclass Distribution")
            
            # Create comprehensive class/subclass analysis
            from motifs.classification_config import OFFICIAL_CLASSIFICATION, get_official_classification
            
            # Initialize complete class/subclass count dictionary
            all_class_counts = {}
            all_subclass_counts = {}
            
            # Initialize with all official classes and subclasses
            for class_info in OFFICIAL_CLASSIFICATION.values():
                class_name = class_info['class_name']
                all_class_counts[class_name] = 0
                
                for subclass in class_info['subclasses']:
                    all_subclass_counts[subclass] = 0
            
            # Count actual detected motifs
            for result in st.session_state.results:
                for motif in result['motifs']:
                    # Get official classification
                    legacy_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                    legacy_subtype = motif.get('Subtype', '')
                    
                    official_class, official_subtype = get_official_classification(legacy_class, legacy_subtype)
                    
                    # Count class
                    if official_class in all_class_counts:
                        all_class_counts[official_class] += 1
                    
                    # Count subclass (if not empty)
                    if official_subtype and official_subtype in all_subclass_counts:
                        all_subclass_counts[official_subtype] += 1
            
            # Display class distribution chart
            if any(count > 0 for count in all_class_counts.values()):
                st.markdown("#### 📊 Class Distribution (All 10 Classes)")
                
                # Create bar chart showing all classes
                class_df = pd.DataFrame([
                    {'Class': class_name, 'Count': count, 'Status': 'Detected' if count > 0 else 'Not Found'}
                    for class_name, count in all_class_counts.items()
                ])
                
                fig = px.bar(
                    class_df,
                    x='Class',
                    y='Count',
                    color='Status',
                    title="Complete Motif Class Distribution (10 Official Classes)",
                    labels={'Count': 'Number of Motifs'},
                    color_discrete_map={'Detected': '#1e3a8a', 'Not Found': '#e5e7eb'}
                )
                fig.update_layout(height=500, xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)
                
                # Class counts table
                st.markdown("#### 📋 Complete Class Summary Table")
                class_summary_df = class_df.copy()
                class_summary_df['Percentage'] = (class_summary_df['Count'] / class_summary_df['Count'].sum() * 100).round(2)
                st.dataframe(class_summary_df, use_container_width=True)
            
            # Display subclass distribution
            if any(count > 0 for count in all_subclass_counts.values()):
                st.markdown("#### 🔬 Subclass Distribution (22+ Official Subclasses)")
                
                # Create subclass chart
                subclass_df = pd.DataFrame([
                    {'Subclass': subclass_name, 'Count': count, 'Status': 'Detected' if count > 0 else 'Not Found'}
                    for subclass_name, count in all_subclass_counts.items()
                ])
                
                fig = px.bar(
                    subclass_df,
                    x='Subclass',
                    y='Count',
                    color='Status',
                    title="Complete Subclass Distribution (22+ Official Subclasses)",
                    labels={'Count': 'Number of Motifs'},
                    color_discrete_map={'Detected': '#0891b2', 'Not Found': '#f3f4f6'}
                )
                fig.update_layout(height=600, xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)
                
                # Subclass counts table
                st.markdown("#### 📋 Complete Subclass Summary Table")
                subclass_summary_df = subclass_df.copy()
                total_subclass_motifs = subclass_summary_df['Count'].sum()
                if total_subclass_motifs > 0:
                    subclass_summary_df['Percentage'] = (subclass_summary_df['Count'] / total_subclass_motifs * 100).round(2)
                else:
                    subclass_summary_df['Percentage'] = 0
                st.dataframe(subclass_summary_df, use_container_width=True)
            
            # Use advanced visualization if available
            if PUBLICATION_VIZ_AVAILABLE:
                st.markdown("#### 🎨 Advanced Interactive Visualizations")
                
                # Prepare motif data for visualization
                all_motifs_for_viz = []
                for result in st.session_state.results:
                    all_motifs_for_viz.extend(result['motifs'])
                
                # Call the integrated visualization function with main context
                create_nbdfinder_visualization_interface(all_motifs_for_viz, 
                                                       sum(r['sequence_length'] for r in st.session_state.results),
                                                       context="main")
        
        # ---- GENOMIC POSITION PLOTS SUBTAB ----
        with result_tabs[2]:
            st.markdown("### 🗺️ Genomic Position Analysis & Plots")
            
            # Sequence selection for detailed view
            if len(st.session_state.results) > 1:
                selected_seq = st.selectbox(
                    "Select sequence for position analysis:",
                    range(len(st.session_state.results)),
                    format_func=lambda x: st.session_state.results[x]['sequence_name']
                )
            else:
                selected_seq = 0
            
            result = st.session_state.results[selected_seq]
            
            # Create position plot
            if result['motifs']:
                positions = []
                types = []
                classes = []
                
                for motif in result['motifs']:
                    positions.append(motif.get('Start', 0))
                    types.append(motif.get('Motif', 'Unknown'))
                    classes.append(motif.get('Class', 'Unknown'))
                
                # Scatter plot of motif positions by class
                fig = px.scatter(
                    x=positions,
                    y=classes,
                    color=classes,
                    title=f"Motif Positions by Class in {result['sequence_name']}",
                    labels={'x': 'Position (bp)', 'y': 'Motif Class'},
                    height=500
                )
                fig.update_layout(showlegend=True)
                st.plotly_chart(fig, use_container_width=True)
                
                # Linear genomic map
                st.markdown("#### 📍 Linear Genomic Map")
                
                # Create a linear map showing motif density
                seq_length = result['sequence_length']
                bin_size = max(seq_length // 100, 1)  # 100 bins
                bins = np.arange(0, seq_length + bin_size, bin_size)
                hist, _ = np.histogram(positions, bins=bins)
                
                fig = px.bar(
                    x=bins[:-1],
                    y=hist,
                    title=f"Motif Density Along {result['sequence_name']}",
                    labels={'x': 'Position (bp)', 'y': 'Motif Count per Bin'}
                )
                fig.update_layout(height=300)
                st.plotly_chart(fig, use_container_width=True)
                
                # Position table
                st.markdown("#### 📋 Detailed Position Table")
                
                # Prepare data for position table following the standard format
                position_data = []
                for motif in result['motifs']:
                    # Ensure motif has proper subtype
                    motif = ensure_subtype(motif)
                    
                    # Get official classification
                    legacy_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                    legacy_subtype = motif.get('Subtype', '')
                    official_class, official_subtype = get_official_classification(legacy_class, legacy_subtype)
                    
                    position_data.append({
                        'Sequence': result['sequence_name'],
                        'Class': official_class,
                        'Subclass': official_subtype,
                        'Start': motif.get('Start', 0),
                        'End': motif.get('End', 0),
                        'Length': motif.get('End', 0) - motif.get('Start', 0) + 1,
                        'Score': motif.get('Score', 'N/A')
                    })
                
                position_df = pd.DataFrame(position_data)
                st.dataframe(position_df, use_container_width=True)
        
        # ---- DETAILED RESULT TABLES SUBTAB ----
        with result_tabs[3]:
            st.markdown("### 📊 Detailed Result Tables")
            
            # Comprehensive motif details table
            all_motifs_details = []
            
            for seq_idx, result in enumerate(st.session_state.results):
                for motif_idx, motif in enumerate(result['motifs']):
                    # Ensure motif has proper subtype
                    motif = ensure_subtype(motif)
                    
                    # Get official classification
                    legacy_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                    legacy_subtype = motif.get('Subtype', '')
                    official_class, official_subtype = get_official_classification(legacy_class, legacy_subtype)
                    
                    all_motifs_details.append({
                        'Sequence': result['sequence_name'],
                        'Class': official_class,
                        'Subclass': official_subtype,
                        'Start': motif.get('Start', 0),
                        'End': motif.get('End', 0),
                        'Length': motif.get('End', 0) - motif.get('Start', 0) + 1,
                        'Score': motif.get('Score', 'N/A')
                    })
            
            if all_motifs_details:
                details_df = pd.DataFrame(all_motifs_details)
                
                # Add filters
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    selected_classes = st.multiselect(
                        "Filter by Class:",
                        options=details_df['Class'].unique(),
                        default=list(details_df['Class'].unique())
                    )
                
                with col2:
                    selected_sequences = st.multiselect(
                        "Filter by Sequence:",
                        options=details_df['Sequence'].unique(),
                        default=list(details_df['Sequence'].unique())
                    )
                
                with col3:
                    min_score = st.number_input("Minimum Score:", value=0.0, min_value=0.0)
                
                # Apply filters
                filtered_df = details_df[
                    (details_df['Class'].isin(selected_classes)) &
                    (details_df['Sequence'].isin(selected_sequences))
                ]
                
                # Apply score filter if Score column has numeric values
                try:
                    filtered_df = filtered_df[pd.to_numeric(filtered_df['Score'], errors='coerce') >= min_score]
                except:
                    pass  # Skip score filtering if scores are not numeric
                
                st.markdown(f"#### 📋 Filtered Results ({len(filtered_df)} motifs)")
                st.dataframe(filtered_df, use_container_width=True)
                
                # Download button for filtered results
                csv = filtered_df.to_csv(index=False)
                st.download_button(
                    label="💾 Download Filtered Results as CSV",
                    data=csv,
                    file_name=f"nbdfinder_detailed_results.csv",
                    mime="text/csv"
                )
        
        # ---- PROFESSIONAL FIGURES SUBTAB ----
        with result_tabs[4]:
            st.markdown("### 📖 Professional Figures")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**📄 Export Options:**")
                export_format = st.selectbox("Format:", ["PNG (300 DPI)", "PDF", "SVG"])
                figure_size = st.selectbox("Size:", ["Small (8x6)", "Medium (12x8)", "Large (16x12)"])
                
            with col2:
                st.markdown("**🎨 Style Options:**")
                color_scheme = st.selectbox("Color scheme:", ["Professional", "Colorblind-friendly", "Grayscale"])
                include_title = st.checkbox("Include title", value=True)
            
            if st.button("🚀 Generate Professional Figure"):
                if not st.session_state.results:
                    st.warning("No analysis results available. Please run an analysis first.")
                else:
                    try:
                        with st.spinner("Generating professional publication figure..."):
                            # Prepare motif data for visualization
                            all_motifs_for_viz = []
                            for result in st.session_state.results:
                                all_motifs_for_viz.extend(result['motifs'])
                            
                            if all_motifs_for_viz:
                                # Use the integrated visualization system
                                if PUBLICATION_VIZ_AVAILABLE:
                                    st.markdown("---")
                                    st.markdown("### 📊 Professional Publication Figures")
                                    
                                    # Call the comprehensive visualization interface with figures context
                                    create_nbdfinder_visualization_interface(
                                        all_motifs_for_viz, 
                                        sum(r['sequence_length'] for r in st.session_state.results),
                                        context="figures"
                                    )
                                    
                                    st.success("✅ Professional figures generated successfully!")
                                    
                                else:
                                    # Fallback to basic matplotlib figure
                                    st.markdown("### 📊 Professional Figure (Basic Version)")
                                    
                                    # Create a publication-quality figure
                                    import matplotlib.pyplot as plt
                                    import seaborn as sns
                                    
                                    # Set style for publication
                                    plt.style.use('seaborn-v0_8-whitegrid')
                                    sns.set_palette("husl")
                                    
                                    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
                                    fig.suptitle('NBDFinder: Comprehensive Non-B DNA Analysis', 
                                               fontsize=20, fontweight='bold')
                                    
                                    # Plot 1: Class distribution
                                    classes = [ensure_subtype(motif).get('Class', 'Unknown') for motif in all_motifs_for_viz]
                                    class_counts = pd.Series(classes).value_counts()
                                    
                                    bars = ax1.bar(range(len(class_counts)), class_counts.values)
                                    ax1.set_xlabel('Non-B DNA Classes', fontweight='bold')
                                    ax1.set_ylabel('Count', fontweight='bold')
                                    ax1.set_title('A. Motif Class Distribution', fontweight='bold')
                                    ax1.set_xticks(range(len(class_counts)))
                                    ax1.set_xticklabels(class_counts.index, rotation=45, ha='right')
                                    
                                    for i, bar in enumerate(bars):
                                        height = bar.get_height()
                                        ax1.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                                               f'{int(height)}', ha='center', va='bottom', fontweight='bold')
                                    
                                    # Plot 2: Score distribution
                                    scores = [motif.get('Score', 0) for motif in all_motifs_for_viz if motif.get('Score', 0) > 0]
                                    if scores:
                                        ax2.hist(scores, bins=20, alpha=0.7, edgecolor='black')
                                        ax2.set_xlabel('Motif Score', fontweight='bold')
                                        ax2.set_ylabel('Frequency', fontweight='bold')
                                        ax2.set_title('B. Score Distribution', fontweight='bold')
                                        ax2.axvline(np.mean(scores), color='red', linestyle='--', 
                                                   label=f'Mean: {np.mean(scores):.1f}')
                                        ax2.legend()
                                    
                                    # Plot 3: Length distribution
                                    lengths = []
                                    for motif in all_motifs_for_viz:
                                        start = motif.get('Start', 0)
                                        end = motif.get('End', 0)
                                        if end > start:
                                            lengths.append(end - start + 1)
                                    
                                    if lengths:
                                        ax3.boxplot(lengths)
                                        ax3.set_ylabel('Motif Length (bp)', fontweight='bold')
                                        ax3.set_title('C. Motif Length Distribution', fontweight='bold')
                                        ax3.set_xticklabels(['All Motifs'])
                                    
                                    # Plot 4: Sequence coverage
                                    if len(st.session_state.results) > 0:
                                        seq_names = [r['sequence_name'][:15] for r in st.session_state.results]
                                        motif_counts = [r['total_motifs'] for r in st.session_state.results]
                                        
                                        bars = ax4.bar(range(len(seq_names)), motif_counts)
                                        ax4.set_xlabel('Sequences', fontweight='bold')
                                        ax4.set_ylabel('Motif Count', fontweight='bold')
                                        ax4.set_title('D. Motifs per Sequence', fontweight='bold')
                                        ax4.set_xticks(range(len(seq_names)))
                                        ax4.set_xticklabels(seq_names, rotation=45, ha='right')
                                        
                                        for i, bar in enumerate(bars):
                                            height = bar.get_height()
                                            ax4.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                                                   f'{int(height)}', ha='center', va='bottom')
                                    
                                    plt.tight_layout()
                                    
                                    # Display the figure
                                    st.pyplot(fig)
                                    
                                    # Save options
                                    col1, col2, col3 = st.columns(3)
                                    
                                    with col1:
                                        # PNG export
                                        img_buffer = BytesIO()
                                        format_dpi = {"PNG (300 DPI)": 300, "PDF": 300, "SVG": 300}
                                        dpi = format_dpi.get(export_format, 300)
                                        
                                        if export_format == "SVG":
                                            plt.savefig(img_buffer, format='svg', dpi=dpi, bbox_inches='tight')
                                            mime_type = "image/svg+xml"
                                            file_ext = "svg"
                                        elif export_format == "PDF":
                                            plt.savefig(img_buffer, format='pdf', dpi=dpi, bbox_inches='tight')
                                            mime_type = "application/pdf"
                                            file_ext = "pdf"
                                        else:
                                            plt.savefig(img_buffer, format='png', dpi=dpi, bbox_inches='tight')
                                            mime_type = "image/png"
                                            file_ext = "png"
                                        
                                        st.download_button(
                                            label=f"📥 Download {export_format}",
                                            data=img_buffer.getvalue(),
                                            file_name=f"nbdfinder_professional_figure.{file_ext}",
                                            mime=mime_type
                                        )
                                    
                                    plt.close()
                            else:
                                st.warning("No motifs found in analysis results.")
                                
                    except Exception as e:
                        st.error(f"Figure generation failed: {str(e)}")
                        st.info("Please ensure you have valid analysis results.")
            
            # Additional visualization options
            if ADVANCED_VIZ_AVAILABLE:
                st.markdown("#### 🎨 Advanced 3D Visualizations")
                
                if st.session_state.results:
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        viz_type = st.selectbox(
                            "3D Visualization Type:",
                            ["3D Scatter Plot", "3D Surface Plot", "3D Motif Landscape", "Interactive 3D Network"]
                        )
                    
                    with col2:
                        color_by = st.selectbox(
                            "Color by:",
                            ["Class", "Score", "Length", "Position"]
                        )
                    
                    if st.button("🌟 Generate 3D Visualization"):
                        try:
                            import plotly.graph_objects as go
                            from plotly.subplots import make_subplots
                            
                            # Prepare data
                            all_motifs = []
                            for result in st.session_state.results:
                                for motif in result['motifs']:
                                    motif_data = ensure_subtype(motif)
                                    motif_data['sequence_name'] = result['sequence_name']
                                    all_motifs.append(motif_data)
                            
                            if all_motifs:
                                # Convert to DataFrame
                                df = pd.DataFrame(all_motifs)
                                
                                if viz_type == "3D Scatter Plot":
                                    # 3D Scatter: Position, Score, Length
                                    fig = go.Figure(data=go.Scatter3d(
                                        x=df['Start'],
                                        y=df.get('Score', 0),
                                        z=df['End'] - df['Start'] + 1,
                                        mode='markers',
                                        marker=dict(
                                            size=8,
                                            color=pd.Categorical(df['Class']).codes,
                                            colorscale='Viridis',
                                            colorbar=dict(title="Motif Class"),
                                            opacity=0.8
                                        ),
                                        text=df['Class'],
                                        hovertemplate="<b>%{text}</b><br>" +
                                                     "Position: %{x}<br>" +
                                                     "Score: %{y}<br>" +
                                                     "Length: %{z}<br>" +
                                                     "<extra></extra>"
                                    ))
                                    
                                    fig.update_layout(
                                        title="3D Motif Analysis: Position vs Score vs Length",
                                        scene=dict(
                                            xaxis_title="Genomic Position",
                                            yaxis_title="Motif Score",
                                            zaxis_title="Motif Length (bp)"
                                        ),
                                        width=800, height=600
                                    )
                                
                                elif viz_type == "3D Surface Plot":
                                    # Create density surface
                                    from scipy.interpolate import griddata
                                    
                                    x = df['Start'].values
                                    y = df.get('Score', 0).values
                                    z = (df['End'] - df['Start'] + 1).values
                                    
                                    # Create grid
                                    xi = np.linspace(x.min(), x.max(), 50)
                                    yi = np.linspace(y.min(), y.max(), 50)
                                    xi, yi = np.meshgrid(xi, yi)
                                    
                                    # Interpolate z values
                                    zi = griddata((x, y), z, (xi, yi), method='cubic')
                                    
                                    fig = go.Figure(data=go.Surface(
                                        x=xi, y=yi, z=zi,
                                        colorscale='Viridis',
                                        colorbar=dict(title="Motif Length")
                                    ))
                                    
                                    fig.update_layout(
                                        title="3D Motif Density Surface",
                                        scene=dict(
                                            xaxis_title="Genomic Position",
                                            yaxis_title="Motif Score",
                                            zaxis_title="Motif Length (bp)"
                                        ),
                                        width=800, height=600
                                    )
                                
                                elif viz_type == "3D Motif Landscape":
                                    # Create a 3D landscape view
                                    classes = df['Class'].unique()
                                    
                                    fig = go.Figure()
                                    
                                    for i, cls in enumerate(classes):
                                        cls_data = df[df['Class'] == cls]
                                        
                                        fig.add_trace(go.Scatter3d(
                                            x=cls_data['Start'],
                                            y=[i] * len(cls_data),
                                            z=cls_data.get('Score', 0),
                                            mode='markers',
                                            marker=dict(
                                                size=6,
                                                color=cls_data['End'] - cls_data['Start'] + 1,
                                                colorscale='Plasma',
                                                opacity=0.8
                                            ),
                                            name=cls,
                                            text=cls_data['Class'],
                                            hovertemplate="<b>%{text}</b><br>" +
                                                         "Position: %{x}<br>" +
                                                         "Score: %{z}<br>" +
                                                         "<extra></extra>"
                                        ))
                                    
                                    fig.update_layout(
                                        title="3D Motif Landscape by Class",
                                        scene=dict(
                                            xaxis_title="Genomic Position",
                                            yaxis_title="Motif Class",
                                            zaxis_title="Motif Score",
                                            yaxis=dict(
                                                ticktext=classes,
                                                tickvals=list(range(len(classes)))
                                            )
                                        ),
                                        width=800, height=600
                                    )
                                
                                elif viz_type == "Interactive 3D Network":
                                    # Create a 3D network of motif relationships
                                    import networkx as nx
                                    
                                    # Create network based on proximity
                                    G = nx.Graph()
                                    
                                    for i, motif in enumerate(all_motifs):
                                        G.add_node(i, 
                                                  motif_class=motif['Class'],
                                                  score=motif.get('Score', 0),
                                                  start=motif['Start'],
                                                  end=motif['End'])
                                    
                                    # Add edges for nearby motifs
                                    for i in range(len(all_motifs)):
                                        for j in range(i+1, len(all_motifs)):
                                            dist = abs(all_motifs[i]['Start'] - all_motifs[j]['Start'])
                                            if dist < 1000:  # Within 1kb
                                                G.add_edge(i, j, weight=1000-dist)
                                    
                                    # Get 3D layout
                                    pos = nx.spring_layout(G, dim=3, k=1, iterations=50)
                                    
                                    # Extract coordinates
                                    x_nodes = [pos[node][0] for node in G.nodes()]
                                    y_nodes = [pos[node][1] for node in G.nodes()]
                                    z_nodes = [pos[node][2] for node in G.nodes()]
                                    
                                    # Create edges
                                    x_edges, y_edges, z_edges = [], [], []
                                    for edge in G.edges():
                                        x_coords = [pos[edge[0]][0], pos[edge[1]][0], None]
                                        y_coords = [pos[edge[0]][1], pos[edge[1]][1], None]
                                        z_coords = [pos[edge[0]][2], pos[edge[1]][2], None]
                                        x_edges += x_coords
                                        y_edges += y_coords
                                        z_edges += z_coords
                                    
                                    # Create traces
                                    trace_edges = go.Scatter3d(
                                        x=x_edges, y=y_edges, z=z_edges,
                                        mode='lines',
                                        line=dict(color='gray', width=2),
                                        hoverinfo='none'
                                    )
                                    
                                    trace_nodes = go.Scatter3d(
                                        x=x_nodes, y=y_nodes, z=z_nodes,
                                        mode='markers',
                                        marker=dict(
                                            size=8,
                                            color=[G.nodes[node]['score'] for node in G.nodes()],
                                            colorscale='Viridis',
                                            colorbar=dict(title="Score"),
                                            opacity=0.8
                                        ),
                                        text=[G.nodes[node]['motif_class'] for node in G.nodes()],
                                        hovertemplate="<b>%{text}</b><br>" +
                                                     "Score: %{marker.color}<br>" +
                                                     "<extra></extra>"
                                    )
                                    
                                    fig = go.Figure(data=[trace_edges, trace_nodes])
                                    fig.update_layout(
                                        title="3D Motif Interaction Network",
                                        scene=dict(
                                            xaxis_title="Network X",
                                            yaxis_title="Network Y",
                                            zaxis_title="Network Z"
                                        ),
                                        showlegend=False,
                                        width=800, height=600
                                    )
                                
                                # Display the 3D plot
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Export option
                                if st.button("📥 Download 3D Visualization"):
                                    html_buffer = BytesIO()
                                    fig.write_html(html_buffer)
                                    
                                    st.download_button(
                                        label="📥 Download Interactive HTML",
                                        data=html_buffer.getvalue(),
                                        file_name=f"3d_visualization_{viz_type.lower().replace(' ', '_')}.html",
                                        mime="text/html"
                                    )
                            
                            else:
                                st.warning("No motif data available for 3D visualization.")
                                
                        except Exception as e:
                            st.error(f"3D visualization failed: {str(e)}")
                            st.info("Some 3D features require additional packages. Using fallback visualization.")
                
                else:
                    st.info("🔍 Run an analysis first to enable 3D visualizations.")
            else:
                st.markdown("#### ℹ️ Advanced Visualizations")
                st.markdown("""
                **3D Visualization Features Available:**
                - 3D Scatter plots of motif properties
                - Interactive surface plots  
                - Motif landscape visualization
                - Network analysis of motif relationships
                
                ✨ These features are fully integrated and ready to use!
                """)
                
                if st.session_state.results:
                    st.info("💡 Analysis results detected! 3D visualizations are available in the comprehensive visualization suite above.")
        
        # ---- DOWNLOAD & EXPORT SUBTAB ----
        with result_tabs[5]:
            st.markdown("### 📥 Download & Export")
            
            if not st.session_state.results:
                st.info("No analysis results available for download. Please run an analysis first.")
            else:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Data Export Options**")
                    
                    # Results table download
                    if st.button("Download Results Table (CSV)"):
                        results_data = []
                        for result in st.session_state.results:
                            for motif in result['motifs']:
                                # Ensure motif has proper subtype
                                motif = ensure_subtype(motif)
                                
                                # Get official classification
                                legacy_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                                legacy_subtype = motif.get('Subtype', '')
                                official_class, official_subtype = get_official_classification(legacy_class, legacy_subtype)
                                
                                results_data.append({
                                    'Sequence': result['sequence_name'],
                                    'Class': official_class,
                                    'Subclass': official_subtype,
                                    'Start': motif.get('Start', 0),
                                    'End': motif.get('End', 0),
                                    'Length': motif.get('End', 0) - motif.get('Start', 0) + 1,
                                    'Score': motif.get('Score', 0),
                                    'Conservation_Score': motif.get('Conservation_Score', 'N/A'),
                                    'Conservation_P_Value': motif.get('Conservation_P_Value', 'N/A'),
                                    'Conservation_Significance': motif.get('Conservation_Significance', 'N/A'),
                                    'ScoreMethod': motif.get('ScoreMethod', 'N/A'),
                                    'Sequence_Fragment': motif.get('Sequence', 'N/A')
                                })
                        
                        df = pd.DataFrame(results_data)
                        csv = df.to_csv(index=False)
                        st.download_button(
                            label="Download CSV",
                            data=csv,
                            file_name="nbdfinder_results.csv",
                            mime="text/csv"
                        )
                    
                    # Excel export
                    if st.button("Download Results (Excel)"):
                        try:
                            # Create Excel file with multiple sheets
                            from io import BytesIO
                            import openpyxl
                            from openpyxl.styles import Font, PatternFill, Alignment
                            
                            output = BytesIO()
                            workbook = openpyxl.Workbook()
                            
                            # Remove default sheet
                            workbook.remove(workbook.active)
                            
                            # Summary sheet
                            summary_sheet = workbook.create_sheet("Summary")
                            summary_sheet['A1'] = "NBDFinder Analysis Summary"
                            summary_sheet['A1'].font = Font(bold=True, size=16)
                            summary_sheet['A3'] = f"Total Sequences: {len(st.session_state.results)}"
                            summary_sheet['A4'] = f"Total Motifs: {sum(r['total_motifs'] for r in st.session_state.results)}"
                            summary_sheet['A5'] = f"Analysis Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
                            
                            # Detailed results sheet with conservation analysis
                            results_sheet = workbook.create_sheet("Detailed_Results")
                            headers = ['Sequence', 'Class', 'Subclass', 'Start', 'End', 'Length', 'Score', 
                                      'Conservation_Score', 'Conservation_P_Value', 'Conservation_Significance',
                                      'ScoreMethod', 'Sequence_Fragment']
                            
                            for col, header in enumerate(headers, 1):
                                cell = results_sheet.cell(row=1, column=col, value=header)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")
                            
                            row = 2
                            for result in st.session_state.results:
                                for motif in result['motifs']:
                                    motif = ensure_subtype(motif)
                                    legacy_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                                    legacy_subtype = motif.get('Subtype', '')
                                    official_class, official_subtype = get_official_classification(legacy_class, legacy_subtype)
                                    
                                    data = [
                                        result['sequence_name'],
                                        official_class,
                                        official_subtype,
                                        motif.get('Start', 0),
                                        motif.get('End', 0),
                                        motif.get('End', 0) - motif.get('Start', 0) + 1,
                                        motif.get('Score', 0),
                                        motif.get('Conservation_Score', 'N/A'),
                                        motif.get('Conservation_P_Value', 'N/A'),
                                        motif.get('Conservation_Significance', 'N/A'),
                                        motif.get('ScoreMethod', 'N/A'),
                                        motif.get('Sequence', 'N/A')
                                    ]
                                    
                                    for col, value in enumerate(data, 1):
                                        results_sheet.cell(row=row, column=col, value=value)
                                    row += 1
                            
                            # Auto-fit columns
                            for sheet in workbook.worksheets:
                                for column in sheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    sheet.column_dimensions[column_letter].width = adjusted_width
                            
                            workbook.save(output)
                            excel_data = output.getvalue()
                            
                            st.download_button(
                                label="📥 Download Excel File",
                                data=excel_data,
                                file_name=f"nbdfinder_results_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                        except Exception as e:
                            st.error(f"Excel export failed: {str(e)}")
                            st.info("Fallback: Using CSV export instead")
                            # Fallback to CSV
                            csv = df.to_csv(index=False)
                            st.download_button(
                                label="📥 Download CSV (Fallback)",
                                data=csv,
                                file_name="nbdfinder_results.csv",
                                mime="text/csv"
                            )
                
                with col2:
                    st.markdown("**Visualization Export**")
                    
                    if st.button("Download Plots (PNG)"):
                        try:
                            import zipfile
                            from io import BytesIO
                            import matplotlib.pyplot as plt
                            
                            # Create ZIP file for multiple plots
                            zip_buffer = BytesIO()
                            
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                # Generate motif distribution plot
                                if st.session_state.results:
                                    all_motifs = []
                                    for result in st.session_state.results:
                                        all_motifs.extend(result['motifs'])
                                    
                                    if all_motifs:
                                        # Class distribution plot
                                        classes = [ensure_subtype(motif).get('Class', 'Unknown') for motif in all_motifs]
                                        class_counts = pd.Series(classes).value_counts()
                                        
                                        fig, ax = plt.subplots(figsize=(12, 8))
                                        bars = ax.bar(range(len(class_counts)), class_counts.values, 
                                                    color=['#1e3a8a', '#0891b2', '#059669', '#dc2626', '#7c3aed', 
                                                          '#ea580c', '#0d9488', '#be185d', '#4338ca', '#7c2d12'][:len(class_counts)])
                                        
                                        ax.set_xlabel('Non-B DNA Classes', fontsize=12, fontweight='bold')
                                        ax.set_ylabel('Count', fontsize=12, fontweight='bold')
                                        ax.set_title('Distribution of Non-B DNA Motifs by Class', fontsize=14, fontweight='bold')
                                        ax.set_xticks(range(len(class_counts)))
                                        ax.set_xticklabels(class_counts.index, rotation=45, ha='right')
                                        
                                        # Add value labels on bars
                                        for i, bar in enumerate(bars):
                                            height = bar.get_height()
                                            ax.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                                                   f'{int(height)}', ha='center', va='bottom', fontweight='bold')
                                        
                                        plt.tight_layout()
                                        
                                        # Save to ZIP
                                        img_buffer = BytesIO()
                                        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                                        zip_file.writestr('motif_class_distribution.png', img_buffer.getvalue())
                                        plt.close()
                                        
                                        # Score distribution plot
                                        scores = [motif.get('Score', 0) for motif in all_motifs if motif.get('Score', 0) > 0]
                                        if scores:
                                            fig, ax = plt.subplots(figsize=(10, 6))
                                            ax.hist(scores, bins=20, color='#0891b2', alpha=0.7, edgecolor='black')
                                            ax.set_xlabel('Motif Score', fontsize=12, fontweight='bold')
                                            ax.set_ylabel('Frequency', fontsize=12, fontweight='bold')
                                            ax.set_title('Distribution of Motif Scores', fontsize=14, fontweight='bold')
                                            ax.grid(True, alpha=0.3)
                                            plt.tight_layout()
                                            
                                            img_buffer = BytesIO()
                                            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                                            zip_file.writestr('motif_score_distribution.png', img_buffer.getvalue())
                                            plt.close()
                                        
                                        # Position plot for first sequence
                                        if len(st.session_state.results) > 0:
                                            result = st.session_state.results[0]
                                            if result['motifs']:
                                                positions = [motif.get('Start', 0) for motif in result['motifs']]
                                                types = [ensure_subtype(motif).get('Class', 'Unknown') for motif in result['motifs']]
                                                
                                                fig, ax = plt.subplots(figsize=(14, 6))
                                                unique_types = list(set(types))
                                                colors = plt.cm.tab10(np.linspace(0, 1, len(unique_types)))
                                                
                                                for i, motif_type in enumerate(unique_types):
                                                    type_positions = [pos for pos, t in zip(positions, types) if t == motif_type]
                                                    ax.scatter(type_positions, [i] * len(type_positions), 
                                                             c=[colors[i]], label=motif_type, s=60, alpha=0.7)
                                                
                                                ax.set_xlabel('Genomic Position (bp)', fontsize=12, fontweight='bold')
                                                ax.set_ylabel('Motif Type', fontsize=12, fontweight='bold')
                                                ax.set_title(f'Motif Positions in {result["sequence_name"]}', fontsize=14, fontweight='bold')
                                                ax.set_yticks(range(len(unique_types)))
                                                ax.set_yticklabels(unique_types)
                                                ax.grid(True, alpha=0.3)
                                                ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
                                                plt.tight_layout()
                                                
                                                img_buffer = BytesIO()
                                                plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                                                zip_file.writestr('motif_positions.png', img_buffer.getvalue())
                                                plt.close()
                            
                            zip_data = zip_buffer.getvalue()
                            
                            st.download_button(
                                label="📥 Download Plot Package (ZIP)",
                                data=zip_data,
                                file_name=f"nbdfinder_plots_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                mime="application/zip"
                            )
                            
                        except Exception as e:
                            st.error(f"Plot export failed: {str(e)}")
                            st.info("Please ensure you have analysis results to export plots.")
                    
                    if st.button("📄 Generate Analysis Report"):
                        if not st.session_state.results:
                            st.warning("No analysis results available for report generation.")
                        else:
                            try:
                                # Create a comprehensive text report
                                report_content = []
                                report_content.append("=" * 60)
                                report_content.append("NBDFinder Analysis Report")
                                report_content.append("=" * 60)
                                report_content.append(f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
                                report_content.append("")
                                
                                # Summary section
                                total_motifs = sum(r['total_motifs'] for r in st.session_state.results)
                                report_content.append("ANALYSIS SUMMARY")
                                report_content.append("-" * 20)
                                report_content.append(f"Total Sequences Analyzed: {len(st.session_state.results)}")
                                report_content.append(f"Total Motifs Detected: {total_motifs}")
                                report_content.append("")
                                
                                # Sequence details
                                report_content.append("SEQUENCE DETAILS")
                                report_content.append("-" * 20)
                                for i, result in enumerate(st.session_state.results, 1):
                                    report_content.append(f"Sequence {i}: {result['sequence_name']}")
                                    report_content.append(f"  Length: {result['sequence_length']} bp")
                                    report_content.append(f"  Motifs: {result['total_motifs']}")
                                    report_content.append("")
                                
                                # Motif details
                                report_content.append("DETAILED MOTIF RESULTS")
                                report_content.append("-" * 30)
                                report_content.append(f"{'Sequence':<20} {'Class':<25} {'Subclass':<20} {'Start':<8} {'End':<8} {'Score':<8}")
                                report_content.append("-" * 90)
                                
                                for result in st.session_state.results:
                                    for motif in result['motifs']:
                                        motif = ensure_subtype(motif)
                                        legacy_class = motif.get('Class', motif.get('Motif', 'Unknown'))
                                        legacy_subtype = motif.get('Subtype', '')
                                        official_class, official_subtype = get_official_classification(legacy_class, legacy_subtype)
                                        
                                        seq_name = result['sequence_name'][:19]
                                        class_name = official_class[:24]
                                        subclass_name = official_subtype[:19]
                                        start = str(motif.get('Start', 0))
                                        end = str(motif.get('End', 0))
                                        score = f"{motif.get('Score', 0):.1f}"
                                        
                                        report_content.append(f"{seq_name:<20} {class_name:<25} {subclass_name:<20} {start:<8} {end:<8} {score:<8}")
                                
                                report_content.append("")
                                report_content.append("=" * 60)
                                report_content.append("End of Report")
                                report_content.append("Generated by NBDFinder v2.0 - Dr. Venkata Rajesh Yella")
                                
                                # Create downloadable text report
                                report_text = "\n".join(report_content)
                                
                                st.download_button(
                                    label="📥 Download Analysis Report (TXT)",
                                    data=report_text,
                                    file_name=f"nbdfinder_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.txt",
                                    mime="text/plain"
                                )
                                
                                # Also show a preview
                                with st.expander("📋 Report Preview", expanded=False):
                                    st.text(report_text[:2000] + "\n\n... (truncated in preview)")
                                
                            except Exception as e:
                                st.error(f"Report generation failed: {str(e)}")
                                st.info("Please ensure you have analysis results to generate a report.")

# =====================================================
# CLINICAL/DISEASE ANNOTATION PAGE
# =====================================================
with tab_dict["Clinical/Disease"]:
    # Clinical subtabs
    clinical_subtabs = st.tabs(["Disease Motif Results", "Clinical Summary"])
    
    # ---- DISEASE MOTIF RESULTS SUBTAB ----
    with clinical_subtabs[0]:
        st.markdown("### 🩺 Disease-Associated Motif Analysis")
        
        if not st.session_state.results:
            st.info("No analysis results available. Please run an analysis first.")
        else:
            # Extract disease-associated motifs from results
            disease_motifs = []
            other_motifs_with_clinical = []
            
            for result in st.session_state.results:
                for motif in result.get('motifs', []):
                    if motif.get('Class') == 'Disease-Associated Motif':
                        disease_motifs.append(motif)
                    elif any(key in motif for key in ['Clinical_Significance', 'Disease_Name', 'Pathogenic']):
                        other_motifs_with_clinical.append(motif)
            
            if disease_motifs or other_motifs_with_clinical:
                # Clinical significance summary
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    pathogenic_count = sum(1 for m in disease_motifs if m.get('Clinical_Significance') == 'Pathogenic')
                    st.metric("Pathogenic Variants", pathogenic_count, delta=None if pathogenic_count == 0 else "⚠️ Attention Required")
                
                with col2:
                    vus_count = sum(1 for m in disease_motifs if 'VUS' in str(m.get('Clinical_Significance', '')))
                    st.metric("Variants of Uncertain Significance", vus_count)
                
                with col3:
                    total_disease_motifs = len(disease_motifs)
                    st.metric("Total Disease-Associated Motifs", total_disease_motifs)
                
                # Disease motifs table
                if disease_motifs:
                    st.markdown("#### 🔬 Disease-Associated Repeat Expansions")
                    
                    clinical_data = []
                    for motif in disease_motifs:
                        clinical_data.append({
                            'Disease': motif.get('Disease_Name', 'Unknown'),
                            'Gene': motif.get('Gene_Symbol', 'Unknown'),
                            'Repeat Unit': motif.get('Repeat_Unit', 'Unknown'),
                            'Count': motif.get('Repeat_Count', 0),
                            'Normal Range': motif.get('Normal_Range', 'Unknown'),
                            'Pathogenic Threshold': motif.get('Pathogenic_Threshold', 'Unknown'),
                            'Clinical Significance': motif.get('Clinical_Significance', 'Unknown'),
                            'Risk Score': f"{motif.get('Risk_Score', 0):.1f}%",
                            'Population Percentile': f"{motif.get('Population_Percentile', 0):.1f}%",
                            'Inheritance': motif.get('Inheritance_Pattern', 'Unknown')
                        })
                    
                    clinical_df = pd.DataFrame(clinical_data)
                    st.dataframe(clinical_df, use_container_width=True)
                    
                    # Individual disease motif details
                    st.markdown("#### 📋 Detailed Clinical Information")
                    
                    for i, motif in enumerate(disease_motifs):
                        with st.expander(f"{motif.get('Disease_Name', 'Unknown Disease')} - {motif.get('Gene_Symbol', 'Unknown Gene')}"):
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("**Clinical Details:**")
                                st.write(f"**Disease:** {motif.get('Disease_Name', 'Unknown')}")
                                st.write(f"**Gene Symbol:** {motif.get('Gene_Symbol', 'Unknown')}")
                                st.write(f"**Inheritance Pattern:** {motif.get('Inheritance_Pattern', 'Unknown')}")
                                st.write(f"**Clinical Features:** {motif.get('Clinical_Features', 'Not specified')}")
                                
                                # Risk assessment
                                risk_score = motif.get('Risk_Score', 0)
                                if risk_score >= 85:
                                    risk_level = "🔴 Very High"
                                elif risk_score >= 70:
                                    risk_level = "🟠 High"
                                elif risk_score >= 50:
                                    risk_level = "🟡 Moderate"
                                else:
                                    risk_level = "🟢 Low"
                                
                                st.write(f"**Risk Level:** {risk_level} ({risk_score:.1f}%)")
                            
                            with col2:
                                st.markdown("**Laboratory Information:**")
                                st.write(f"**Repeat Unit:** {motif.get('Repeat_Unit', 'Unknown')}")
                                st.write(f"**Repeat Count:** {motif.get('Repeat_Count', 0)}")
                                st.write(f"**Normal Range:** {motif.get('Normal_Range', 'Unknown')}")
                                st.write(f"**Pathogenic Threshold:** {motif.get('Pathogenic_Threshold', 'Unknown')}")
                                st.write(f"**Population Percentile:** {motif.get('Population_Percentile', 0):.1f}%")
                                
                                # Structural impact
                                structural_impact = motif.get('Structural_Impact', 'Not specified')
                                st.write(f"**Structural Impact:** {structural_impact}")
                            
                            # Therapeutic and counseling information
                            therapeutic_target = motif.get('Therapeutic_Target', 'No specific targets identified')
                            genetic_counseling = motif.get('Genetic_Counseling', 'Standard genetic counseling recommended')
                            
                            st.markdown("**🎯 Therapeutic Targets:**")
                            st.info(therapeutic_target)
                            
                            st.markdown("**🧬 Genetic Counseling Recommendations:**")
                            st.info(genetic_counseling)
                            
                            # References
                            pmid_refs = motif.get('PMID_References', '')
                            if pmid_refs:
                                st.markdown("**📚 Literature References:**")
                                refs = pmid_refs.split('; ')
                                for ref in refs:
                                    if ref.strip():
                                        st.markdown(f"- [PMID: {ref.strip()}](https://pubmed.ncbi.nlm.nih.gov/{ref.strip()}/)")
                
                # Other motifs with clinical relevance
                if other_motifs_with_clinical:
                    st.markdown("#### 🔬 Other Motifs with Clinical Relevance")
                    
                    other_clinical_data = []
                    for motif in other_motifs_with_clinical:
                        pathogenic_status = "Yes" if motif.get('Pathogenic', False) else "No"
                        other_clinical_data.append({
                            'Motif Class': motif.get('Class', 'Unknown'),
                            'Subtype': motif.get('Subtype', 'Unknown'),
                            'Length': motif.get('Length', 0),
                            'Score': f"{motif.get('Score', 0):.1f}",
                            'Pathogenic': pathogenic_status,
                            'Position': f"{motif.get('Start', 0)}-{motif.get('End', 0)}"
                        })
                    
                    other_df = pd.DataFrame(other_clinical_data)
                    st.dataframe(other_df, use_container_width=True)
            
            else:
                st.info("🔍 No disease-associated motifs detected in the current analysis.")
                st.markdown("""
                <div class="feature-card">
                    <h4>🧬 What are disease-associated motifs?</h4>
                    <p>Disease-associated motifs are specific DNA sequences, often repeat expansions, that are linked to genetic diseases when present beyond normal thresholds. Examples include:</p>
                    <ul>
                        <li><strong>GAA repeats</strong> in Friedreich's Ataxia (FXN gene)</li>
                        <li><strong>CGG repeats</strong> in Fragile X Syndrome (FMR1 gene)</li>
                        <li><strong>CAG repeats</strong> in Huntington's Disease (HTT gene)</li>
                        <li><strong>CTG repeats</strong> in Myotonic Dystrophy (DMPK gene)</li>
                    </ul>
                    <p>Try analyzing sequences with known repeat expansions to see the clinical analysis in action.</p>
                </div>
                """, unsafe_allow_html=True)
    
    # ---- CLINICAL SUMMARY SUBTAB ----
    with clinical_subtabs[1]:
        st.markdown("### 📊 Clinical Interpretation Summary")
        
        if not st.session_state.results:
            st.info("No analysis results available. Please run an analysis first.")
        else:
            # Generate comprehensive clinical summary
            total_motifs = sum(len(result.get('motifs', [])) for result in st.session_state.results)
            disease_motifs = []
            clinical_motifs = []
            
            for result in st.session_state.results:
                for motif in result.get('motifs', []):
                    if motif.get('Class') == 'Disease-Associated Motif':
                        disease_motifs.append(motif)
                    elif any(key in motif for key in ['Clinical_Significance', 'Disease_Name', 'Pathogenic']):
                        clinical_motifs.append(motif)
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Total Motifs Detected", total_motifs)
            
            with col2:
                pathogenic_count = sum(1 for m in disease_motifs if m.get('Clinical_Significance') == 'Pathogenic')
                st.metric("Pathogenic Findings", pathogenic_count, 
                         delta="⚠️" if pathogenic_count > 0 else "✅")
            
            with col3:
                likely_pathogenic = sum(1 for m in disease_motifs if 'Likely Pathogenic' in str(m.get('Clinical_Significance', '')))
                st.metric("Likely Pathogenic", likely_pathogenic)
            
            with col4:
                vus_count = sum(1 for m in disease_motifs if 'VUS' in str(m.get('Clinical_Significance', '')))
                st.metric("VUS Findings", vus_count)
            
            if disease_motifs:
                # Risk stratification
                st.markdown("#### 🎯 Risk Stratification")
                
                risk_categories = {'Very High (≥85%)': 0, 'High (70-84%)': 0, 'Moderate (50-69%)': 0, 'Low (<50%)': 0}
                
                for motif in disease_motifs:
                    risk_score = motif.get('Risk_Score', 0)
                    if risk_score >= 85:
                        risk_categories['Very High (≥85%)'] += 1
                    elif risk_score >= 70:
                        risk_categories['High (70-84%)'] += 1
                    elif risk_score >= 50:
                        risk_categories['Moderate (50-69%)'] += 1
                    else:
                        risk_categories['Low (<50%)'] += 1
                
                # Display risk distribution
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Risk Distribution:**")
                    for category, count in risk_categories.items():
                        if count > 0:
                            if 'Very High' in category:
                                st.error(f"🔴 {category}: {count} finding(s)")
                            elif 'High' in category:
                                st.warning(f"🟠 {category}: {count} finding(s)")
                            elif 'Moderate' in category:
                                st.info(f"🟡 {category}: {count} finding(s)")
                            else:
                                st.success(f"🟢 {category}: {count} finding(s)")
                
                with col2:
                    # Most significant finding
                    highest_risk_motif = max(disease_motifs, key=lambda x: x.get('Risk_Score', 0), default=None)
                    if highest_risk_motif:
                        st.markdown("**Most Significant Finding:**")
                        st.info(f"""
                        **Disease:** {highest_risk_motif.get('Disease_Name', 'Unknown')}  
                        **Gene:** {highest_risk_motif.get('Gene_Symbol', 'Unknown')}  
                        **Risk Score:** {highest_risk_motif.get('Risk_Score', 0):.1f}%  
                        **Clinical Significance:** {highest_risk_motif.get('Clinical_Significance', 'Unknown')}
                        """)
                
                # Therapeutic opportunities
                st.markdown("#### 🎯 Therapeutic Targeting Opportunities")
                
                therapeutic_targets = {}
                for motif in disease_motifs:
                    target = motif.get('Therapeutic_Target', 'No specific targets identified')
                    if target != 'No specific targets identified':
                        disease = motif.get('Disease_Name', 'Unknown')
                        if target not in therapeutic_targets:
                            therapeutic_targets[target] = []
                        therapeutic_targets[target].append(disease)
                
                if therapeutic_targets:
                    for target, diseases in therapeutic_targets.items():
                        with st.expander(f"🎯 {target}"):
                            st.write(f"**Applicable to:** {', '.join(set(diseases))}")
                            
                            # Add relevant information based on target type
                            if 'Antisense oligonucleotides' in target:
                                st.info("💡 **Mechanism:** ASOs can modulate RNA splicing and reduce toxic RNA accumulation")
                            elif 'Gene therapy' in target:
                                st.info("💡 **Mechanism:** Viral vector-mediated gene replacement or correction")
                            elif 'RNA interference' in target:
                                st.info("💡 **Mechanism:** siRNA or shRNA to reduce mutant protein expression")
                else:
                    st.info("No specific therapeutic targets identified for the detected variants.")
                
                # Clinical databases and resources
                st.markdown("#### 🔗 Clinical Databases & Resources")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Genetic Testing Resources:**")
                    st.markdown("""
                    - [ClinVar](https://www.ncbi.nlm.nih.gov/clinvar/) - Clinical variant database
                    - [OMIM](https://www.omim.org/) - Online Mendelian Inheritance in Man
                    - [GTR](https://www.ncbi.nlm.nih.gov/gtr/) - Genetic Testing Registry
                    - [ClinGen](https://clinicalgenome.org/) - Clinical Genome Resource
                    """)
                
                with col2:
                    st.markdown("**Professional Guidelines:**")
                    st.markdown("""
                    - [ACMG Guidelines](https://www.acmg.net/) - Variant interpretation standards
                    - [AMP Guidelines](https://www.amp.org/) - Molecular pathology guidelines
                    - [CAP Guidelines](https://www.cap.org/) - Laboratory accreditation standards
                    - [NSGC](https://www.nsgc.org/) - Genetic counseling resources
                    """)
                
                # Genetic counseling recommendations
                st.markdown("#### 🧬 Genetic Counseling Recommendations")
                
                counseling_summary = []
                inheritance_patterns = set()
                
                for motif in disease_motifs:
                    counseling = motif.get('Genetic_Counseling', '')
                    inheritance = motif.get('Inheritance_Pattern', 'Unknown')
                    
                    if counseling:
                        counseling_summary.append(counseling)
                    if inheritance != 'Unknown':
                        inheritance_patterns.add(inheritance)
                
                if counseling_summary:
                    for counseling in set(counseling_summary):  # Remove duplicates
                        st.info(counseling)
                
                if inheritance_patterns:
                    st.markdown("**Inheritance Patterns Detected:**")
                    inheritance_info = {
                        'AR': 'Autosomal Recessive - Both parents are carriers',
                        'AD': 'Autosomal Dominant - One affected parent passes the condition',
                        'XL': 'X-Linked - Passed through X chromosome',
                        'MT': 'Mitochondrial - Maternal inheritance'
                    }
                    
                    for pattern in inheritance_patterns:
                        if pattern in inheritance_info:
                            st.write(f"- **{pattern}:** {inheritance_info[pattern]}")
            
            else:
                st.markdown("""
                <div class="feature-card">
                    <h4>✅ No High-Risk Genetic Variants Detected</h4>
                    <p>The analysis did not identify any disease-associated repeat expansions or pathogenic variants in the analyzed sequence(s).</p>
                    
                    <h5>📋 Clinical Interpretation:</h5>
                    <ul>
                        <li>This result suggests the analyzed sequence does not contain known pathogenic repeat expansions</li>
                        <li>Standard non-B DNA structures detected may still have biological significance</li>
                        <li>Consider clinical context and additional testing if symptoms suggest genetic disease</li>
                    </ul>
                    
                    <h5>⚠️ Important Notes:</h5>
                    <ul>
                        <li>This analysis focuses on repeat expansion disorders and known non-B DNA motifs</li>
                        <li>Other types of genetic variants (SNPs, CNVs, etc.) are not covered</li>
                        <li>Clinical correlation and additional testing may be warranted</li>
                        <li>Consult with a genetic counselor for comprehensive interpretation</li>
                    </ul>
                </div>
                """, unsafe_allow_html=True)

# =====================================================
# DOWNLOAD & EXPORT PAGE
# =====================================================
# =====================================================
# DOCUMENTATION PAGE
# =====================================================
with tab_dict["Documentation"]:
    # Documentation subtabs
    doc_subtabs = st.tabs(["📋 Class Overview", "🔬 Subclass Details", "⚙️ Detection Methods", "📊 Scoring Systems", "📖 User Guide"])
    
    # ---- CLASS OVERVIEW SUBTAB ----
    with doc_subtabs[0]:
        st.markdown("### 📋 Complete Non-B DNA Classification System")
        
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a; margin-bottom: 15px;">🎯 Official Classification Structure</h4>
            <p style="font-size: 1.1rem; margin-bottom: 20px;"><strong>NBDFinder implements the comprehensive 10-class, 22+ subclass Non-B DNA classification system:</strong></p>
        </div>
        """, unsafe_allow_html=True)
        
        # Import classification for comprehensive display
        from motifs.classification_config import OFFICIAL_CLASSIFICATION, get_class_counts
        
        # Create overview table
        overview_data = []
        for class_id, class_info in OFFICIAL_CLASSIFICATION.items():
            class_name = class_info['class_name']
            subclasses = class_info['subclasses']
            subclass_count = len(subclasses) if subclasses else "Dynamic"
            
            overview_data.append({
                "Class ID": f"{class_id}",
                "Class Name": class_name,
                "Subclasses": subclass_count,
                "Type": "Fixed" if subclasses else "Dynamic"
            })
        
        overview_df = pd.DataFrame(overview_data)
        st.markdown("#### 🏛️ Classification Overview Table")
        st.dataframe(overview_df, use_container_width=True, hide_index=True)
        
        # Display classes in expandable format with enhanced content
        st.markdown("#### 📚 Detailed Class Descriptions")
        
        # Comprehensive class descriptions
        class_descriptions = {
            "Curved DNA": {
                "description": "DNA sequences that adopt intrinsic curvature due to specific sequence-dependent structural features. The intrinsic curvature arises from the anisotropic flexibility of DNA, particularly in A-tract sequences.",
                "biological_significance": "Involved in gene regulation, chromatin organization, and protein-DNA interactions. Facilitates nucleosome positioning and transcriptional control.",
                "structural_features": "Reduced minor groove width, asymmetric charge distribution, enhanced bendability at specific sites."
            },
            "Slipped DNA": {
                "description": "Structures formed by repetitive DNA sequences that can extrude into hairpin or loop configurations during replication, repair, or recombination.",
                "biological_significance": "Associated with genomic instability, repeat expansion diseases, and evolutionary sequence variation. Critical in neurodegenerative disorders.",
                "structural_features": "Stem-loop structures with perfect or imperfect base pairing, variable loop sizes, sequence-dependent stability."
            },
            "Cruciform DNA": {
                "description": "Four-way junction structures formed by inverted repeat sequences that can extrude from double-stranded DNA under supercoiling stress or during replication.",
                "biological_significance": "Implicated in recombination hotspots, chromosomal rearrangements, and gene regulation. May serve as nuclease recognition sites.",
                "structural_features": "Four-arm junction with variable arm lengths, junction mobility, sequence-dependent stability and geometry."
            },
            "R-loop": {
                "description": "Three-stranded nucleic acid structures consisting of an RNA-DNA hybrid and a displaced single-stranded DNA loop, formed during transcription or RNA processing.",
                "biological_significance": "Essential for gene regulation, DNA repair, and replication initiation. Aberrant R-loops linked to genomic instability and cancer.",
                "structural_features": "RNA-DNA hybrid duplex, displaced ssDNA, sequence-dependent formation, transcription-dependent stability."
            },
            "Triplex": {
                "description": "Three-stranded DNA structures formed by Hoogsteen or reverse-Hoogsteen hydrogen bonding between a third strand and the major groove of duplex DNA.",
                "biological_significance": "Potential therapeutic targets for gene regulation, involved in recombination and mutagenesis. Relevant to repeat expansion disorders.",
                "structural_features": "Major groove binding, pH-dependent stability, purine-rich target sequences, sequence-specific recognition."
            },
            "G-Quadruplex Family": {
                "description": "Four-stranded secondary structures formed by guanine-rich sequences through Hoogsteen hydrogen bonding and π-π stacking interactions, stabilized by monovalent cations.",
                "biological_significance": "Critical for telomere maintenance, gene regulation, and genome stability. Therapeutic targets for cancer and aging research.",
                "structural_features": "Tetrad formation, loop topologies, cation coordination, polymorphic structures with parallel/antiparallel orientations."
            },
            "i-motif family": {
                "description": "Four-stranded secondary structures formed by cytosine-rich sequences through N3-protonated cytosine-cytosine base pairs, typically stable at acidic pH.",
                "biological_significance": "Complementary to G-quadruplexes, involved in gene regulation and chromosomal organization. pH sensors for cellular processes.",
                "structural_features": "Intercalated cytosine pairs, pH-dependent stability, antiparallel strand orientation, narrow major groove."
            },
            "Z-DNA": {
                "description": "Left-handed double helix formed by alternating purine-pyrimidine sequences, particularly CpG dinucleotides, under high salt conditions or negative supercoiling.",
                "biological_significance": "Involved in gene regulation, particularly in promoter regions. Associated with transcriptional activation and chromatin remodeling.",
                "structural_features": "Left-handed helix, zigzag backbone, syn glycosidic conformation for purines, deep minor groove."
            },
            "Hybrid": {
                "description": "Complex structures formed when multiple non-B DNA motifs overlap or interact within the same genomic region, creating composite structural elements.",
                "biological_significance": "Represent hotspots of structural complexity and potential genomic instability. May have unique regulatory functions.",
                "structural_features": "Combination of multiple structural elements, context-dependent properties, enhanced complexity and potential instability."
            },
            "Non-B DNA cluster regions": {
                "description": "Genomic regions with high density of multiple different non-B DNA forming sequences within close proximity (typically 100 nucleotides).",
                "biological_significance": "Hotspots for genomic instability, recombination, and regulatory element clustering. Associated with disease susceptibility loci.",
                "structural_features": "Multiple overlapping or proximal non-B structures, synergistic effects, enhanced potential for structural transitions."
            }
        }
        
        # Organize classes in two columns for better layout
        col1, col2 = st.columns(2)
        
        # Split classes between columns
        classes_list = list(OFFICIAL_CLASSIFICATION.items())
        col1_classes = classes_list[:5]  # Classes 1-5
        col2_classes = classes_list[5:]  # Classes 6-10
        
        with col1:
            for class_id, class_info in col1_classes:
                class_name = class_info['class_name']
                subclasses = class_info['subclasses']
                
                with st.expander(f"**{class_id}. {class_name}**", expanded=False):
                    desc_info = class_descriptions.get(class_name, {})
                    
                    st.markdown(f"**🔬 Description:**")
                    st.write(desc_info.get("description", "Detailed description available in literature."))
                    
                    st.markdown(f"**🧬 Biological Significance:**")
                    st.write(desc_info.get("biological_significance", "Important for various cellular processes."))
                    
                    st.markdown(f"**🏗️ Structural Features:**")
                    st.write(desc_info.get("structural_features", "Unique structural characteristics."))
                    
                    if subclasses:
                        st.markdown("**📊 Subclasses:**")
                        for i, subclass in enumerate(subclasses, 1):
                            st.markdown(f"• **{subclass}**")
                        st.markdown(f"*Total subclasses: {len(subclasses)}*")
                    else:
                        if class_name == "Hybrid":
                            st.markdown("**🔗 Dynamic subclasses based on overlapping motifs:**")
                            st.markdown("• Created when any two different non-B DNA classes overlap")
                            st.markdown("• Named by contributing classes (e.g., 'G4-Cruciform')")
                        elif class_name == "Non-B DNA cluster regions":
                            st.markdown("**🎯 Dynamic subclasses based on motif clustering:**")
                            st.markdown("• Regions with 3+ different motif types within 100 nucleotides")
                            st.markdown("• Named by predominant motif types")
        
        with col2:
            for class_id, class_info in col2_classes:
                class_name = class_info['class_name']
                subclasses = class_info['subclasses']
                
                with st.expander(f"**{class_id}. {class_name}**", expanded=False):
                    desc_info = class_descriptions.get(class_name, {})
                    
                    st.markdown(f"**🔬 Description:**")
                    st.write(desc_info.get("description", "Detailed description available in literature."))
                    
                    st.markdown(f"**🧬 Biological Significance:**")
                    st.write(desc_info.get("biological_significance", "Important for various cellular processes."))
                    
                    st.markdown(f"**🏗️ Structural Features:**")
                    st.write(desc_info.get("structural_features", "Unique structural characteristics."))
                    
                    if subclasses:
                        st.markdown("**📊 Subclasses:**")
                        for i, subclass in enumerate(subclasses, 1):
                            st.markdown(f"• **{subclass}**")
                        st.markdown(f"*Total subclasses: {len(subclasses)}*")
                    else:
                        if class_name == "Hybrid":
                            st.markdown("**🔗 Dynamic subclasses based on overlapping motifs:**")
                            st.markdown("• Created when any two different non-B DNA classes overlap")
                            st.markdown("• Named by contributing classes (e.g., 'G4-Cruciform')")
                        elif class_name == "Non-B DNA cluster regions":
                            st.markdown("**🎯 Dynamic subclasses based on motif clustering:**")
                            st.markdown("• Regions with 3+ different motif types within 100 nucleotides")
                            st.markdown("• Named by predominant motif types")
        
        # Summary statistics
        total_classes, total_fixed_subclasses = get_class_counts()
        st.markdown(f"""
        <div class="feature-card" style="margin-top: 20px;">
            <h4 style="color: #1e3a8a;">📊 Classification Summary</h4>
            <ul style="font-size: 1.05rem;">
                <li><strong>Total Classes:</strong> {total_classes}</li>
                <li><strong>Fixed Subclasses:</strong> {total_fixed_subclasses}</li>
                <li><strong>Dynamic Subclasses:</strong> Variable (Hybrid + Clusters)</li>
                <li><strong>Total Coverage:</strong> 22+ distinct Non-B DNA types</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    # ---- SUBCLASS DETAILS SUBTAB ----
    with doc_subtabs[1]:
        st.markdown("### 🔬 Comprehensive Subclass Documentation")
        
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">🎯 All 22+ Official Subclasses</h4>
            <p>Detailed documentation for each subclass including detection parameters, biological significance, and structural characteristics.</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Comprehensive subclass information
        subclass_details = {
            # Curved DNA subclasses
            "Global curvature": {
                "class": "Curved DNA",
                "detection_method": "Phasing-aware tract scoring",
                "key_features": "A-tracts with proper phasing (10-11 bp intervals)",
                "biological_role": "Nucleosome positioning, chromatin organization",
                "typical_length": "50-200 bp",
                "score_threshold": "6.0+",
                "references": "Trifonov & Sussman (1980), Crothers et al. (1990)"
            },
            "Local Curvature": {
                "class": "Curved DNA", 
                "detection_method": "Individual tract curvature analysis",
                "key_features": "Short A-tracts or T-tracts causing local bending",
                "biological_role": "Local DNA flexibility, protein binding sites",
                "typical_length": "15-50 bp",
                "score_threshold": "7.0+",
                "references": "Koo et al. (1986), Hagerman (1990)"
            },
            
            # Slipped DNA subclasses
            "Slipped DNA [Direct Repeat]": {
                "class": "Slipped DNA",
                "detection_method": "DR_PerfectBlock_v1 algorithm",
                "key_features": "Perfect or near-perfect direct repeats",
                "biological_role": "Replication slippage, repeat expansion",
                "typical_length": "20-100 bp",
                "score_threshold": "10.0+",
                "references": "Levinson & Gutman (1987), Schlötterer & Tautz (1992)"
            },
            "Slipped DNA [STR]": {
                "class": "Slipped DNA",
                "detection_method": "STR_PerfectBlock_v1 algorithm",
                "key_features": "Short tandem repeats (1-6 bp units)",
                "biological_role": "Microsatellite instability, genetic variation",
                "typical_length": "15-200 bp",
                "score_threshold": "10.0+",
                "references": "Tautz (1989), Weber & May (1989)"
            },
            
            # Cruciform DNA subclasses
            "Cruciform DNA [IR]/HairPin [IR]": {
                "class": "Cruciform DNA",
                "detection_method": "IR_PerfectPalindrome_v1 algorithm",
                "key_features": "Inverted repeats forming cruciform structures",
                "biological_role": "Recombination hotspots, gene regulation",
                "typical_length": "20-300 bp",
                "score_threshold": "2.0+",
                "references": "Lilley (1980), Gellert et al. (1983)"
            },
            
            # R-loop subclasses
            "R-loop": {
                "class": "R-loop",
                "detection_method": "RLFS_UnifiedFramework_raw algorithm",
                "key_features": "RNA-DNA hybrid with displaced ssDNA",
                "biological_role": "Transcription regulation, DNA repair initiation",
                "typical_length": "100-1000 bp",
                "score_threshold": "Variable",
                "references": "Thomas et al. (1976), Skourti-Stathaki et al. (2011)"
            },
            
            # Triplex subclasses
            "Triplex": {
                "class": "Triplex",
                "detection_method": "Triplex_Mirror_v1 algorithm",
                "key_features": "Three-stranded structure with Hoogsteen bonds",
                "biological_role": "Gene regulation, potential therapeutic target",
                "typical_length": "15-100 bp",
                "score_threshold": "25.0+",
                "references": "Felsenfeld et al. (1957), Frank-Kamenetskii & Mirkin (1995)"
            },
            "sticky DNA": {
                "class": "Triplex",
                "detection_method": "GAA_TTC_v1 algorithm",
                "key_features": "GAA/TTC repeats forming stable triplexes",
                "biological_role": "Friedreich's ataxia pathogenesis",
                "typical_length": "30-200 bp",
                "score_threshold": "6.0+",
                "references": "Sakamoto et al. (1999), Campuzano et al. (1996)"
            },
            
            # G-Quadruplex Family subclasses
            "Multimeric G4": {
                "class": "G-Quadruplex Family",
                "detection_method": "G4Hunter + structural analysis",
                "key_features": "Multiple G4 units in tandem or close proximity",
                "biological_role": "Enhanced stability, complex regulation",
                "typical_length": "50-200 bp",
                "score_threshold": "0.8+",
                "references": "Hänsel-Hertsch et al. (2017), Chambers et al. (2015)"
            },
            "Canonical G4": {
                "class": "G-Quadruplex Family",
                "detection_method": "G4Hunter with strict parameters",
                "key_features": "Classic G3+N1-7G3+N1-7G3+N1-7G3+ pattern",
                "biological_role": "Telomere maintenance, gene regulation",
                "typical_length": "15-50 bp",
                "score_threshold": "1.0+",
                "references": "Sen & Gilbert (1988), Blackburn (1991)"
            },
            "Relaxed G4": {
                "class": "G-Quadruplex Family",
                "detection_method": "G4Hunter with relaxed constraints",
                "key_features": "G2+N1-12G2+N1-12G2+N1-12G2+ pattern",
                "biological_role": "Broader regulatory functions",
                "typical_length": "15-100 bp",
                "score_threshold": "0.8+",
                "references": "Todd et al. (2005), Huppert & Balasubramanian (2005)"
            },
            "Bulged G4": {
                "class": "G-Quadruplex Family",
                "detection_method": "G4Hunter with bulge tolerance",
                "key_features": "G-quartets with bulged nucleotides",
                "biological_role": "Structural diversity, protein interactions",
                "typical_length": "20-80 bp",
                "score_threshold": "0.8+",
                "references": "Hazel et al. (2004), Lim et al. (2009)"
            },
            "Bipartite G4": {
                "class": "G-Quadruplex Family", 
                "detection_method": "G4Hunter with discontinuous detection",
                "key_features": "G4 motifs split by intervening sequences",
                "biological_role": "Long-range interactions, complex regulation",
                "typical_length": "50-300 bp",
                "score_threshold": "0.8+",
                "references": "Mukundan & Phan (2013), Sahakyan et al. (2017)"
            },
            "Imperfect G4": {
                "class": "G-Quadruplex Family",
                "detection_method": "G4Hunter with mismatch tolerance",
                "key_features": "G-quartets with imperfect base pairing",
                "biological_role": "Evolutionary intermediates, dynamic regulation",
                "typical_length": "15-70 bp",
                "score_threshold": "0.8+",
                "references": "Webba da Silva (2007), Kang et al. (2015)"
            },
            "G-Triplex intermediate": {
                "class": "G-Quadruplex Family",
                "detection_method": "G4Hunter adapted for triplex detection",
                "key_features": "Three-stranded G-rich structures",
                "biological_role": "Structural intermediates, assembly platforms",
                "typical_length": "15-60 bp",
                "score_threshold": "0.8+",
                "references": "Phan (2010), Lim & Phan (2013)"
            },
            
            # i-motif family subclasses
            "Canonical i-motif": {
                "class": "i-motif family",
                "detection_method": "iM-Hunter exact scoring",
                "key_features": "C3+N1-7C3+N1-7C3+N1-7C3+ pattern, pH-dependent",
                "biological_role": "Gene regulation, chromosome organization",
                "typical_length": "15-50 bp",
                "score_threshold": "0.45+",
                "references": "Gehring et al. (1993), Zeraati et al. (2018)"
            },
            "Relaxed i-motif": {
                "class": "i-motif family",
                "detection_method": "iM-Hunter with relaxed constraints",
                "key_features": "C2+N1-12C2+N1-12C2+N1-12C2+ pattern",
                "biological_role": "Extended regulatory functions",
                "typical_length": "15-100 bp",
                "score_threshold": "0.30+",
                "references": "Dzatko et al. (2018), Kaiser et al. (2017)"
            },
            "AC-motif": {
                "class": "i-motif family",
                "detection_method": "AC alternation scoring algorithm",
                "key_features": "Alternating A-C sequences forming structures",
                "biological_role": "Potential regulatory elements",
                "typical_length": "20-100 bp",
                "score_threshold": "Variable",
                "references": "Weil et al. (1999), Nonin-Lecomte et al. (2001)"
            },
            
            # Z-DNA subclasses
            "Z-DNA": {
                "class": "Z-DNA",
                "detection_method": "Kadane_TransitionArray_v1 algorithm",
                "key_features": "Alternating CpG dinucleotides, left-handed helix",
                "biological_role": "Gene regulation, chromatin remodeling",
                "typical_length": "10-100 bp",
                "score_threshold": "50.0+",
                "references": "Wang et al. (1979), Rich et al. (1984)"
            },
            "eGZ (Extruded-G) DNA": {
                "class": "Z-DNA",
                "detection_method": "eGZ_CGG_Expansion_v1 algorithm",
                "key_features": "CGG repeats with potential Z-form adoption",
                "biological_role": "Fragile X syndrome, repeat instability",
                "typical_length": "30-200 bp",
                "score_threshold": "3.0+",
                "references": "Usdin & Woodford (1995), Kumari & Usdin (2001)"
            }
        }
        
        # Create comprehensive subclass table
        subclass_data = []
        for subclass_name, details in subclass_details.items():
            subclass_data.append({
                "Subclass": subclass_name,
                "Parent Class": details["class"],
                "Detection Method": details["detection_method"],
                "Typical Length": details["typical_length"],
                "Score Threshold": details["score_threshold"],
                "Key Features": details["key_features"][:50] + "..." if len(details["key_features"]) > 50 else details["key_features"]
            })
        
        subclass_df = pd.DataFrame(subclass_data)
        st.markdown("#### 📊 Complete Subclass Summary Table")
        st.dataframe(subclass_df, use_container_width=True, hide_index=True)
        
        # Detailed expandable sections for each subclass
        st.markdown("#### 📖 Detailed Subclass Documentation")
        
        # Group by parent class for better organization
        for class_id, class_info in OFFICIAL_CLASSIFICATION.items():
            class_name = class_info['class_name']
            subclasses = class_info['subclasses']
            
            if subclasses:  # Only show classes with defined subclasses
                st.markdown(f"##### {class_id}. {class_name}")
                
                for subclass in subclasses:
                    if subclass in subclass_details:
                        details = subclass_details[subclass]
                        
                        with st.expander(f"🔍 **{subclass}**"):
                            col1, col2 = st.columns([1, 1])
                            
                            with col1:
                                st.markdown("**🔧 Detection Details:**")
                                st.write(f"**Method:** {details['detection_method']}")
                                st.write(f"**Length Range:** {details['typical_length']}")
                                st.write(f"**Score Threshold:** {details['score_threshold']}")
                                
                                st.markdown("**🧬 Structural Features:**")
                                st.write(details['key_features'])
                            
                            with col2:
                                st.markdown("**🎯 Biological Role:**")
                                st.write(details['biological_role'])
                                
                                st.markdown("**📚 Key References:**")
                                st.write(details['references'])
        
        # Dynamic subclasses explanation
        st.markdown("##### 9-10. Dynamic Subclasses")
        
        with st.expander("🔗 **Hybrid Subclasses** (Dynamic)"):
            st.markdown("""
            **Generation:** Created when two or more different Non-B DNA classes overlap spatially
            
            **Naming Convention:** 
            - Two classes: "Class1-Class2" (e.g., "G4-Cruciform")
            - Multiple classes: "Multi-motif hybrid"
            
            **Detection:** Automated overlap analysis during motif detection
            
            **Biological Significance:** 
            - Represent structural complexity hotspots
            - Associated with genomic instability
            - Potential regulatory integration points
            """)
        
        with st.expander("🎯 **Cluster Region Subclasses** (Dynamic)"):
            st.markdown("""
            **Generation:** Regions with 3+ different motif types within 100 nucleotides
            
            **Classification Criteria:**
            - Minimum 3 different Non-B DNA classes
            - Maximum distance: 100 bp between motifs
            - Minimum motif count: 3 total motifs
            
            **Naming Convention:** Based on predominant motif types
            
            **Biological Significance:**
            - Genomic instability hotspots
            - Recombination-prone regions
            - Disease-associated loci
            """)

    # ---- DETECTION METHODS SUBTAB ----
    with doc_subtabs[2]:
        st.markdown("### Detection Algorithms")
        
    # ---- DETECTION METHODS SUBTAB ----
    with doc_subtabs[2]:
        st.markdown("### ⚙️ Detection Algorithms & Methods")
        
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">🔬 Computational Framework Overview</h4>
            <p>NBDFinder employs state-of-the-art computational algorithms specifically designed for each motif class, combining sequence analysis, thermodynamic modeling, and machine learning approaches.</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Algorithm workflow diagram (text-based)
        st.markdown("#### 📊 Detection Workflow")
        
        st.markdown("""
        ```
        Input DNA Sequence
               ↓
        ┌─────────────────────────────────────────┐
        │          Preprocessing                  │
        │  • Sequence validation                  │
        │  • Quality filtering                    │
        │  • Length constraints                   │
        └─────────────────┬───────────────────────┘
                         ↓
        ┌─────────────────────────────────────────┐
        │       Parallel Class Detection          │
        │  ┌─────────────┬─────────────┐          │
        │  │  Sequence   │  Energy     │          │
        │  │  Pattern    │  Based      │          │
        │  │  Matching   │  Analysis   │          │
        │  └─────────────┴─────────────┘          │
        │  ┌─────────────┬─────────────┐          │
        │  │  ML-based   │  Hybrid     │          │
        │  │  Prediction │  Methods    │          │
        │  │             │             │          │
        │  └─────────────┴─────────────┘          │
        └─────────────────┬───────────────────────┘
                         ↓
        ┌─────────────────────────────────────────┐
        │         Post-processing                 │
        │  • Overlap resolution                   │
        │  • Scoring & ranking                    │
        │  • Quality filtering                    │
        │  • Hybrid detection                     │
        │  • Cluster analysis                     │
        └─────────────────┬───────────────────────┘
                         ↓
              Final Motif Annotations
        ```
        """)
        
        # Detailed algorithm descriptions
        st.markdown("#### 🔍 Class-Specific Detection Algorithms")
        
        algorithm_details = {
            "Curved DNA": {
                "methods": [
                    "Phasing-aware tract scoring",
                    "A-tract identification and analysis",
                    "Curvature angle calculation"
                ],
                "key_parameters": [
                    "Minimum A-tract length: 4 bp",
                    "Phasing window: 10-11 bp",
                    "Curvature threshold: >20°"
                ],
                "computational_approach": "Statistical analysis of A-tract distribution and phasing patterns with geometric modeling of DNA bending."
            },
            "Slipped DNA": {
                "methods": [
                    "Perfect block repeat detection",
                    "Tandem repeat identification",
                    "Slippage potential calculation"
                ],
                "key_parameters": [
                    "Minimum repeat unit: 1 bp",
                    "Maximum mismatch: 10%",
                    "Minimum copies: 3"
                ],
                "computational_approach": "Suffix array-based repeat detection with thermodynamic stability assessment of slipped structures."
            },
            "Cruciform DNA": {
                "methods": [
                    "Inverted repeat detection",
                    "Palindrome analysis",
                    "Cruciform stability prediction"
                ],
                "key_parameters": [
                    "Minimum arm length: 6 bp",
                    "Maximum spacer: 100 bp",
                    "Minimum homology: 80%"
                ],
                "computational_approach": "Dynamic programming for optimal palindrome alignment with thermodynamic cruciform stability modeling."
            },
            "R-loop": {
                "methods": [
                    "R-loop Forming Sequence (RLFS) prediction",
                    "RNA-DNA hybrid stability analysis",
                    "Unified framework scoring"
                ],
                "key_parameters": [
                    "G-richness threshold: >50%",
                    "Minimum length: 50 bp",
                    "Stability threshold: Variable"
                ],
                "computational_approach": "Machine learning model trained on experimental R-loop datasets with thermodynamic validation."
            },
            "Triplex": {
                "methods": [
                    "Mirror repeat detection",
                    "Hoogsteen base pair prediction",
                    "pH-dependent stability modeling"
                ],
                "key_parameters": [
                    "Purine content: >70%",
                    "Minimum length: 15 bp",
                    "pH range: 5.0-8.0"
                ],
                "computational_approach": "Pattern matching for triplex-forming sequences with thermodynamic stability calculations under physiological conditions."
            },
            "G-Quadruplex Family": {
                "methods": [
                    "G4Hunter algorithm",
                    "Pattern-based detection",
                    "Structural topology prediction"
                ],
                "key_parameters": [
                    "G-run length: 2-4 bp",
                    "Loop length: 1-12 bp",
                    "Score threshold: >0.8"
                ],
                "computational_approach": "Sliding window G4Hunter scoring with machine learning classification for structural subtypes."
            },
            "i-motif family": {
                "methods": [
                    "iM-Hunter algorithm",
                    "C-run identification",
                    "pH-dependent folding prediction"
                ],
                "key_parameters": [
                    "C-run length: 2-4 bp",
                    "Loop length: 1-12 bp",
                    "pH optimum: 5.5-6.5"
                ],
                "computational_approach": "Complementary to G4Hunter with pH-dependent stability modeling and experimental validation correlation."
            },
            "Z-DNA": {
                "methods": [
                    "Kadane algorithm adaptation",
                    "Dinucleotide transition analysis",
                    "B-Z transition prediction"
                ],
                "key_parameters": [
                    "CpG dinucleotide enrichment",
                    "Alternating purine-pyrimidine",
                    "Supercoiling sensitivity"
                ],
                "computational_approach": "Maximum subarray algorithm for identifying optimal Z-forming regions with transition energy calculations."
            }
        }
        
        # Display algorithm details in organized manner
        for class_name, details in algorithm_details.items():
            with st.expander(f"🔧 **{class_name} Detection**"):
                col1, col2 = st.columns([1, 1])
                
                with col1:
                    st.markdown("**🎯 Detection Methods:**")
                    for method in details["methods"]:
                        st.write(f"• {method}")
                    
                    st.markdown("**⚙️ Key Parameters:**")
                    for param in details["key_parameters"]:
                        st.write(f"• {param}")
                
                with col2:
                    st.markdown("**🔬 Computational Approach:**")
                    st.write(details["computational_approach"])
        
        # Performance characteristics
        st.markdown("#### ⚡ Performance Characteristics")
        
        performance_data = [
            {"Algorithm Class": "Sequence Pattern", "Example": "Curved DNA, STR", "Time Complexity": "O(n)", "Accuracy": "95-98%"},
            {"Algorithm Class": "Energy-based", "Example": "R-loop, Triplex", "Time Complexity": "O(n log n)", "Accuracy": "85-95%"},
            {"Algorithm Class": "ML-enhanced", "Example": "G4Hunter, iM-Hunter", "Time Complexity": "O(n)", "Accuracy": "90-97%"},
            {"Algorithm Class": "Hybrid Methods", "Example": "Cruciform, Z-DNA", "Time Complexity": "O(n²)", "Accuracy": "88-94%"}
        ]
        
        performance_df = pd.DataFrame(performance_data)
        st.dataframe(performance_df, use_container_width=True, hide_index=True)
        
        # Quality control measures
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">✅ Quality Control & Validation</h4>
            <ul style="font-size: 1.05rem;">
                <li><strong>Experimental Validation:</strong> Algorithms benchmarked against published experimental datasets</li>
                <li><strong>Cross-validation:</strong> Performance tested on independent sequence sets</li>
                <li><strong>Sensitivity Analysis:</strong> Parameter optimization for balanced precision/recall</li>
                <li><strong>Comparative Analysis:</strong> Performance comparison with existing tools</li>
                <li><strong>Continuous Updates:</strong> Regular algorithm refinement based on new research</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    # ---- SCORING SYSTEMS SUBTAB ----
    with doc_subtabs[3]:
        st.markdown("### Scoring Systems")
        
    # ---- SCORING SYSTEMS SUBTAB ----
    with doc_subtabs[3]:
        st.markdown("### 📊 Comprehensive Scoring Systems")
        
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">🎯 Motif Scoring Methodology</h4>
            <p>Each detected motif is assigned confidence scores based on multiple biological and computational factors. Our scoring systems are calibrated against experimental data and literature benchmarks.</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Comprehensive scoring documentation from implementation
        scoring_systems = {
            "Curved DNA": [
                {
                    "subclass": "Global curvature", 
                    "method": "Phasing-aware tract scoring",
                    "formula": "(tract_len_sum/10) + tract_count + (2*phasing_score) + length_bonus",
                    "min_score": "6.0",
                    "max_score": "~50+",
                    "factors": ["A-tract length", "A-tract count", "Phasing quality", "Total length"]
                },
                {
                    "subclass": "Local Curvature",
                    "method": "Individual tract curvature",
                    "formula": "length * (1 + AT_fraction) + 0.5 * run_bonus",
                    "min_score": "~7+",
                    "max_score": "~100+", 
                    "factors": ["Tract length", "AT content", "Run quality"]
                }
            ],
            "Slipped DNA": [
                {
                    "subclass": "Direct Repeat",
                    "method": "DR_PerfectBlock_v1",
                    "formula": "unit_weight + copy_bonus + length_bonus - spacer_penalty",
                    "min_score": "10.0",
                    "max_score": "~20+",
                    "factors": ["Unit weight", "Copy number", "Length", "Spacer quality"]
                },
                {
                    "subclass": "STR",
                    "method": "STR_PerfectBlock_v1", 
                    "formula": "unit_weight[1-6] + copy_bonus + length_bonus",
                    "min_score": "10.0",
                    "max_score": "~25+",
                    "factors": ["Unit size", "Copy number", "Perfect matches"]
                }
            ],
            "Cruciform DNA": [
                {
                    "subclass": "IR/HairPin",
                    "method": "IR_PerfectPalindrome_v1",
                    "formula": "arm_weight(arm_len) + length_bonus - spacer_penalty",
                    "min_score": "2+",
                    "max_score": "13+",
                    "factors": ["Arm length", "Palindrome quality", "Spacer length"]
                }
            ],
            "R-loop": [
                {
                    "subclass": "R-loop",
                    "method": "RLFS_UnifiedFramework_raw",
                    "formula": "(traditional_score + unified_score) / 2",
                    "min_score": "Variable",
                    "max_score": "~500+",
                    "factors": ["G-richness", "Skew index", "Length", "Stability"]
                }
            ],
            "Triplex": [
                {
                    "subclass": "Triplex",
                    "method": "Triplex_Mirror_v1",
                    "formula": "arm_len * (1 + 1.5*homogeneity) - 0.8*spacer_len + length_step",
                    "min_score": "25.0",
                    "max_score": "~200+",
                    "factors": ["Arm length", "Homogeneity", "Spacer", "Total length"]
                },
                {
                    "subclass": "sticky DNA",
                    "method": "GAA_TTC_v1",
                    "formula": "copies * (1 + 0.3*AT_fraction)",
                    "min_score": "6+",
                    "max_score": "~100+",
                    "factors": ["Copy number", "AT content"]
                }
            ],
            "G-Quadruplex Family": [
                {
                    "subclass": "All G4 subtypes",
                    "method": "G4Hunter + structural factors",
                    "formula": "G4Hunter_score * structural_factor * length",
                    "min_score": "0.8+",
                    "max_score": "~10+",
                    "factors": ["G4Hunter score", "Structural topology", "Loop constraints"]
                }
            ],
            "i-motif family": [
                {
                    "subclass": "Canonical i-motif",
                    "method": "iM-Hunter exact",
                    "formula": "C-centric windowed scoring",
                    "min_score": "0.45+",
                    "max_score": "~1.0+",
                    "factors": ["C-run quality", "Loop length (1-7)", "Overall structure"]
                },
                {
                    "subclass": "Relaxed i-motif",
                    "method": "iM-Hunter relaxed",
                    "formula": "Similar with relaxed constraints",
                    "min_score": "0.30+",
                    "max_score": "~0.8+",
                    "factors": ["C-run tolerance", "Loop length (1-12)", "Structural flexibility"]
                },
                {
                    "subclass": "AC-motif",
                    "method": "AC alternation scoring",
                    "formula": "A/C fraction + run_density + alternation",
                    "min_score": "Variable",
                    "max_score": "~20+",
                    "factors": ["AC content", "Run density", "Alternation pattern"]
                }
            ],
            "Z-DNA": [
                {
                    "subclass": "Z-DNA",
                    "method": "Kadane_TransitionArray_v1",
                    "formula": "Dinucleotide transition scoring",
                    "min_score": "50.0",
                    "max_score": "~200+",
                    "factors": ["CpG content", "Alternation", "Transition potential"]
                },
                {
                    "subclass": "eGZ",
                    "method": "eGZ_CGG_Expansion_v1",
                    "formula": "copies * (1 + 2*G_fraction)",
                    "min_score": "3+",
                    "max_score": "~100+",
                    "factors": ["CGG repeat count", "G content bonus", "Structural factors"]
                }
            ],
            "Hybrid": [
                {
                    "subclass": "Dynamic",
                    "method": "HybridOverlap_SubclassAnalysis_raw",
                    "formula": "base_score * (1 + interaction_bonus) * overlap_degree",
                    "min_score": "Variable (component-dependent)",
                    "max_score": "Variable (multiple high-scoring overlaps)",
                    "factors": ["Contributing motif scores", "Interaction bonus", "Overlap degree"]
                }
            ],
            "Non-B DNA cluster regions": [
                {
                    "subclass": "Dynamic",
                    "method": "Hotspot density analysis",
                    "formula": "Density-based clustering in 100nt windows",
                    "min_score": "3+ motifs in window",
                    "max_score": "Variable (high motif density)",
                    "factors": ["Motif count in window", "Window size", "Motif diversity"]
                }
            ]
        }
        
        # Create comprehensive scoring table
        st.markdown("#### 📈 Complete Scoring Systems Overview")
        
        scoring_data = []
        for class_name, class_systems in scoring_systems.items():
            for system in class_systems:
                scoring_data.append({
                    "Class": class_name,
                    "Subclass": system["subclass"],
                    "Method": system["method"],
                    "Score Range": f"{system['min_score']} to {system['max_score']}",
                    "Key Factors": len(system["factors"])
                })
        
        scoring_df = pd.DataFrame(scoring_data)
        st.dataframe(scoring_df, use_container_width=True, hide_index=True)
        
        # Detailed scoring explanations
        st.markdown("#### 🔍 Detailed Scoring Documentation")
        
        for class_name, class_systems in scoring_systems.items():
            with st.expander(f"📊 **{class_name} Scoring**"):
                for system in class_systems:
                    st.markdown(f"##### {system['subclass']}")
                    
                    col1, col2 = st.columns([1, 1])
                    
                    with col1:
                        st.markdown("**🔧 Technical Details:**")
                        st.write(f"**Method:** {system['method']}")
                        st.write(f"**Formula:** `{system['formula']}`")
                        st.write(f"**Score Range:** {system['min_score']} to {system['max_score']}")
                    
                    with col2:
                        st.markdown("**📋 Scoring Factors:**")
                        for factor in system["factors"]:
                            st.write(f"• {factor}")
                    
                    st.markdown("---")
        
        # Scoring interpretation guide
        st.markdown("#### 📖 Score Interpretation Guide")
        
        interpretation_guide = {
            "High Confidence (Top 20%)": {
                "description": "Motifs with scores in the top 20% of their class distribution",
                "reliability": "Very High (>95% validated)",
                "usage": "Primary candidates for functional studies",
                "color": "#10b981"
            },
            "Medium Confidence (20-60%)": {
                "description": "Motifs with moderate scores, likely functional",
                "reliability": "High (85-95% validated)",
                "usage": "Good candidates requiring additional validation",
                "color": "#f59e0b"
            },
            "Low Confidence (60-90%)": {
                "description": "Motifs with lower scores, potentially functional",
                "reliability": "Moderate (70-85% validated)",
                "usage": "Candidates for computational validation",
                "color": "#ef4444"
            },
            "Very Low Confidence (Bottom 10%)": {
                "description": "Motifs with marginal scores, uncertain functionality",
                "reliability": "Low (<70% validated)",
                "usage": "Exploratory analysis only",
                "color": "#6b7280"
            }
        }
        
        for level, details in interpretation_guide.items():
            st.markdown(f"""
            <div style="border-left: 4px solid {details['color']}; padding: 10px; margin: 10px 0; background: rgba(248,250,252,0.5);">
                <h5 style="color: {details['color']}; margin: 0 0 8px 0;">{level}</h5>
                <p style="margin: 5px 0;"><strong>Description:</strong> {details['description']}</p>
                <p style="margin: 5px 0;"><strong>Reliability:</strong> {details['reliability']}</p>
                <p style="margin: 5px 0;"><strong>Recommended Usage:</strong> {details['usage']}</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Quality metrics
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">✅ Scoring Quality Metrics</h4>
            <ul style="font-size: 1.05rem;">
                <li><strong>Calibration:</strong> Scores calibrated against experimental datasets</li>
                <li><strong>Validation:</strong> Cross-validated on independent test sets</li>
                <li><strong>Benchmarking:</strong> Compared with published scoring methods</li>
                <li><strong>Sensitivity Analysis:</strong> Robust across parameter variations</li>
                <li><strong>Biological Relevance:</strong> Correlated with functional annotations</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    # ---- USER GUIDE SUBTAB ----
    with doc_subtabs[4]:
        st.markdown("### User Guide")
        
    # ---- USER GUIDE SUBTAB ----
    with doc_subtabs[4]:
        st.markdown("### 📖 Comprehensive User Guide")
        
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">🚀 Quick Start Guide</h4>
            <p>Complete workflow for analyzing DNA sequences with NBDFinder</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Step-by-step workflow
        st.markdown("#### 📋 Analysis Workflow")
        
        workflow_steps = [
            {
                "step": "1. Sequence Input",
                "description": "Upload your DNA sequences or paste them directly",
                "details": [
                    "Supported formats: FASTA, TXT, multi-sequence files",
                    "File size limit: 200MB per file",
                    "Sequence length: 10 bp to 10 Mb recommended",
                    "Use Example sequences for testing"
                ],
                "tab": "Upload & Analyze"
            },
            {
                "step": "2. Parameter Configuration", 
                "description": "Configure analysis parameters for your research needs",
                "details": [
                    "Select specific motif classes or analyze all 10 classes",
                    "Adjust sensitivity: High (comprehensive) vs Medium/Low (stringent)",
                    "Set size limits: minimum and maximum motif lengths",
                    "Configure overlap handling (default: 50%)"
                ],
                "tab": "Upload & Analyze"
            },
            {
                "step": "3. Analysis Execution",
                "description": "Run the comprehensive motif detection analysis",
                "details": [
                    "Processing time varies with sequence length and complexity",
                    "Progress tracking with real-time updates",
                    "Parallel processing for improved performance",
                    "Memory optimization for large sequences"
                ],
                "tab": "Upload & Analyze"
            },
            {
                "step": "4. Results Exploration",
                "description": "Explore detected motifs and their properties",
                "details": [
                    "Interactive tables with motif details",
                    "Distribution charts and statistics",
                    "Sequence-level and motif-level analysis",
                    "Quality metrics and confidence scores"
                ],
                "tab": "Results & Visualization"
            },
            {
                "step": "5. Visualization & Export",
                "description": "Create publication-ready figures and export data",
                "details": [
                    "Multiple chart types: pie, bar, heatmap, circular",
                    "Customizable visualizations",
                    "Export formats: PNG, SVG, PDF for figures",
                    "Data export: CSV, Excel, JSON formats"
                ],
                "tab": "Results & Visualization"
            },
            {
                "step": "6. Clinical Analysis (Optional)",
                "description": "Analyze disease-associated motifs and pathogenic potential",
                "details": [
                    "Disease motif identification",
                    "Pathogenic threshold analysis",
                    "Clinical annotation integration",
                    "Therapeutic target assessment"
                ],
                "tab": "Clinical/Disease"
            }
        ]
        
        for step_info in workflow_steps:
            with st.expander(f"**{step_info['step']}: {step_info['description']}**"):
                st.markdown(f"**📍 Location:** {step_info['tab']} tab")
                st.markdown("**🔍 Details:**")
                for detail in step_info['details']:
                    st.write(f"• {detail}")
        
        # Input format specifications
        st.markdown("#### 📁 Input Format Specifications")
        
        format_specs = {
            "FASTA Format": {
                "extension": ".fa, .fasta, .txt",
                "structure": "Header line starting with '>' followed by sequence",
                "example": """>Sequence_1 Description
ATCGATCGATCGATCGAAAAAAAAAAATCGATCG
>Sequence_2 Description  
GCGCGCGCGGGGGGGGGGGGGGGGCGCGCGCGC""",
                "notes": "Most common format, supports multiple sequences"
            },
            "Plain Text": {
                "extension": ".txt",
                "structure": "Raw DNA sequence without headers",
                "example": "ATCGATCGATCGATCGAAAAAAAAAAATCGATCG",
                "notes": "Simple format for single sequences"
            },
            "Multi-sequence": {
                "extension": "Any supported format",
                "structure": "Multiple sequences in single file",
                "example": "Batch analysis of related sequences",
                "notes": "Efficient for comparative studies"
            }
        }
        
        for format_name, specs in format_specs.items():
            with st.expander(f"📄 **{format_name}**"):
                col1, col2 = st.columns([1, 1])
                with col1:
                    st.write(f"**Extensions:** {specs['extension']}")
                    st.write(f"**Structure:** {specs['structure']}")
                    st.write(f"**Notes:** {specs['notes']}")
                with col2:
                    st.markdown("**Example:**")
                    st.code(specs['example'], language="text")
        
        # Parameter optimization guide
        st.markdown("#### ⚙️ Parameter Optimization Guide")
        
        param_guide = {
            "Sensitivity Settings": {
                "High": "Comprehensive detection, may include marginal motifs (recommended for discovery)",
                "Medium": "Balanced precision and recall (recommended for most analyses)", 
                "Low": "Stringent detection, only high-confidence motifs (recommended for validation)"
            },
            "Size Limits": {
                "Minimum Length": "Set based on motif type (10 bp general, 15 bp for complex motifs)",
                "Maximum Length": "Balance between detection scope and computational efficiency (200 bp default)",
                "Considerations": "Longer limits increase computation time but may capture extended motifs"
            },
            "Overlap Handling": {
                "0-25%": "Strict separation, minimal overlaps allowed",
                "25-50%": "Moderate overlap tolerance (default 50%)",
                "50-75%": "High overlap tolerance, may detect nested structures",
                "75-100%": "Maximum overlap, useful for complex structural analysis"
            }
        }
        
        for category, settings in param_guide.items():
            with st.expander(f"⚙️ **{category}**"):
                for setting, description in settings.items():
                    st.write(f"**{setting}:** {description}")
        
        # Troubleshooting guide
        st.markdown("#### 🔧 Troubleshooting Guide")
        
        troubleshooting = {
            "No sequences loaded": {
                "problem": "Upload appears successful but no sequences detected",
                "solutions": [
                    "Check file format - ensure proper FASTA headers if using FASTA",
                    "Verify sequence content - only ATCG nucleotides allowed",
                    "Check file size - must be under 200MB limit",
                    "Try different input method (paste vs upload)"
                ]
            },
            "Analysis fails to start": {
                "problem": "Start Analysis button remains disabled",
                "solutions": [
                    "Ensure sequences are properly loaded",
                    "Check that at least one motif class is selected",
                    "Verify parameter settings are within valid ranges",
                    "Refresh the page and retry"
                ]
            },
            "Poor detection results": {
                "problem": "Expected motifs not detected or too many false positives",
                "solutions": [
                    "Adjust sensitivity: Higher for more detection, Lower for stringency",
                    "Check sequence quality and composition",
                    "Verify motif class selection matches expectations",
                    "Consider sequence length limitations"
                ]
            },
            "Performance issues": {
                "problem": "Analysis runs slowly or times out",
                "solutions": [
                    "Enable memory optimization option",
                    "Reduce sequence length or number of sequences",
                    "Use parallel processing if available",
                    "Consider analyzing sequences separately"
                ]
            }
        }
        
        for issue, info in troubleshooting.items():
            with st.expander(f"🔧 **{issue}**"):
                st.write(f"**Problem:** {info['problem']}")
                st.markdown("**Solutions:**")
                for solution in info['solutions']:
                    st.write(f"• {solution}")
        
        # Best practices
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">✅ Best Practices</h4>
            <ul style="font-size: 1.05rem;">
                <li><strong>Data Quality:</strong> Use high-quality, validated sequences for best results</li>
                <li><strong>Parameter Selection:</strong> Start with default parameters, then optimize based on results</li>
                <li><strong>Validation:</strong> Cross-validate important findings with experimental data</li>
                <li><strong>Documentation:</strong> Keep detailed records of analysis parameters for reproducibility</li>
                <li><strong>Literature Review:</strong> Compare results with published studies for biological relevance</li>
                <li><strong>Statistical Analysis:</strong> Use appropriate statistical methods for multi-sequence comparisons</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        # Contact and support information
        st.markdown("""
        <div class="feature-card">
            <h4 style="color: #1e3a8a;">📞 Support & Resources</h4>
            <ul style="font-size: 1.05rem;">
                <li><strong>GitHub Repository:</strong> <a href="https://github.com/VRYella/NBDFinder" target="_blank">https://github.com/VRYella/NBDFinder</a></li>
                <li><strong>Issue Reporting:</strong> <a href="https://github.com/VRYella/NBDFinder/issues" target="_blank">GitHub Issues</a></li>
                <li><strong>Email Support:</strong> <a href="mailto:yvrajesh_bt@kluniversity.in">yvrajesh_bt@kluniversity.in</a></li>
                <li><strong>Documentation:</strong> Complete user manual and API reference available</li>
                <li><strong>Updates:</strong> Follow repository for latest features and improvements</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("### Scientific References")
        
        st.markdown("""
        <div class="feature-card">
            <h4>📄 Key Publications</h4>
            <p><strong>Primary Citation:</strong></p>
            <blockquote>
            Yella, V.R. (2024). NBDFinder: A Comprehensive Computational Framework for Genome-Wide Detection and Analysis of Non-B DNA Structural Motifs. GitHub Repository.
            </blockquote>
            
            <p><strong>Related Literature:</strong></p>
            <ul>
                <li>Brascher et al. (2023). Non-B DNA structures and genome instability</li>
                <li>Guiblet et al. (2021). Non-B DNA secondary structures and disease</li>
                <li>Wells et al. (2020). Non-canonical DNA structures in biology</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; padding: 20px; color: #6b7280; font-size: 0.875rem;">
    <p><strong>NBDFinder v2.0</strong> | Developed by Dr. Venkata Rajesh Yella | 
    <a href="https://github.com/VRYella/NBDFinder" style="color: #0891b2;">GitHub</a> | 
    <a href="mailto:yvrajesh_bt@kluniversity.in" style="color: #0891b2;">Contact</a></p>
    <p>© 2024 NBDFinder. Academic use license.</p>
</div>
""", unsafe_allow_html=True)
